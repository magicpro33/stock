# -----------------------------------
# REQUIREMENTS
# pip install streamlit yfinance pandas numpy openpyxl tqdm requests
#
# RUN:
#   streamlit run app.py
# -----------------------------------

import io
import sys
import time
import requests
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.utils import get_column_letter
import yfinance as yf

# ───────────────────────────────────────────────────────────────
# PAGE CONFIG
# ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Stock Screener",
    page_icon="📈",
    layout="wide",
)

st.title("📈 Hybrid Stock Screener")
st.caption("Screens S&P 500, NYSE, or NASDAQ for individual companies — ETFs, funds, and index products excluded.")

tab_screener, tab_analyze = st.tabs(["📊 Screener", "🔍 Analyze a Stock"])

# ───────────────────────────────────────────────────────────────
# METRIC CONFIG  — name, default weight, full description
# ───────────────────────────────────────────────────────────────
METRICS = {
    "OE_Yield": {
        "label":   "OE Yield",
        "weight":  3,
        "desc":    "Owner Earnings Yield = (Net Income + Depreciation − CapEx) ÷ Market Cap. "
                   "Warren Buffett's preferred measure of true cash profitability relative to price. "
                   "Higher is better — above 5% is considered good value.",
    },
    "ROIC": {
        "label":   "ROIC",
        "weight":  2,
        "desc":    "Return on Invested Capital = NOPAT ÷ Invested Capital. "
                   "Measures how efficiently a company generates profit from the capital deployed. "
                   "Above 15% suggests a durable competitive advantage (moat).",
    },
    "ROIC_Trend": {
        "label":   "ROIC Trend",
        "weight":  2,
        "desc":    "Year-over-year change in ROIC. A rising ROIC means the company's moat is "
                   "widening — it's earning more from each dollar of capital over time. "
                   "Positive = improving; negative = deteriorating.",
    },
    "RevenueGrowth": {
        "label":   "Revenue Growth",
        "weight":  1,
        "desc":    "Year-over-year revenue growth rate. Measures top-line expansion. "
                   "Above 5% annually is considered healthy for a mature company; "
                   "high-growth companies often show 15–30%+.",
    },
    "EarningsGrowth": {
        "label":   "Earnings Growth",
        "weight":  1,
        "desc":    "Year-over-year earnings (EPS) growth rate. Confirms that revenue growth "
                   "is translating into real profit. Ideally grows faster than revenue, "
                   "indicating improving operating leverage.",
    },
    "Piotroski": {
        "label":   "Piotroski Score",
        "weight":  1,
        "desc":    "9-point accounting health score (F-Score) developed by Stanford professor "
                   "Joseph Piotroski. Tests profitability, leverage, and operating efficiency. "
                   "7–9 = financially strong; 0–2 = weak or potentially distressed.",
    },
    "OBV": {
        "label":   "OBV (On-Balance Volume)",
        "weight":  1,
        "desc":    "On-Balance Volume tracks cumulative buying vs selling pressure by adding "
                   "volume on up-days and subtracting on down-days. A rising OBV slope over "
                   "the last 20 days means institutions are quietly accumulating shares. "
                   "Score: 1.0 if slope is positive, 0.0 if negative.",
    },
    "MFI": {
        "label":   "MFI (Money Flow Index)",
        "weight":  1,
        "desc":    "Money Flow Index is a volume-weighted RSI (0–100). It measures whether "
                   "money is flowing into or out of a stock. Score = MFI ÷ 100, so it ranges "
                   "0.0–1.0 across the full scale. Below 0.3 = oversold/outflow. "
                   "0.5 = neutral. Above 0.6 = buying pressure. Above 0.8 = strong inflow.",
    },
    "PCV": {
        "label":   "PCV (Price-Confirmed Volume)",
        "weight":  1,
        "desc":    "Price-Confirmed Volume measures what fraction of recent volume occurred "
                   "on days the stock closed higher than the prior day. If >55% of volume "
                   "happens on up-days, buyers are in control. Score scales from 0 (50/50) "
                   "to 1.0 (all up-day volume).",
    },
    "RangePosScore": {
        "label":   "Range Position Score",
        "weight":  2,
        "desc":    "Breakout proximity score — measures how close the current price is to the "
                   "LOW end of its trading range. Score = 1 − RangePos, so 1.0 means price is "
                   "sitting exactly at range support (maximum coiled energy), and 0.0 means "
                   "price is at the top of the range (already extended). Use this with the "
                   "Price Range Filter to find stocks coiling at the bottom of a tight range.",
    },
    "RSI": {
        "label":   "RSI (Momentum Zone)",
        "weight":  2,
        "desc":    "14-day Relative Strength Index scored for continuation potential. "
                   "1.0 = RSI 55–70 (sweet spot: uptrend confirmed, not overbought). "
                   "0.6 = RSI 50–55 (momentum building). "
                   "0.2 = RSI 70–80 (caution: extended). "
                   "0.0 = RSI >80 (overbought/reversal risk) or <50 (no uptrend).",
    },
    "MACD": {
        "label":   "MACD Momentum",
        "weight":  2,
        "desc":    "MACD histogram score — measures short-term momentum acceleration. "
                   "1.0 = histogram positive AND growing (momentum accelerating upward). "
                   "0.6 = histogram positive but shrinking (momentum slowing). "
                   "0.2 = histogram just crossed zero (early signal). "
                   "0.0 = histogram negative (downward momentum). "
                   "MACD line = 12-day EMA minus 26-day EMA. Signal = 9-day EMA of MACD.",
    },
    "GoldenCross": {
        "label":   "Golden Cross",
        "weight":  3,
        "desc":    "Golden Cross = 50-day MA is above the 200-day MA (long-term uptrend confirmed). "
                   "1.0 = 50MA > 200MA (golden cross in effect — institutional buy signal). "
                   "0.5 = 50MA within 2% of 200MA (about to cross — early setup). "
                   "0.0 = 50MA < 200MA (death cross — downtrend). "
                   "Requires at least 200 bars of history.",
    },
    "MFISweetSpot": {
        "label":   "MFI Sweet Spot",
        "weight":  2,
        "desc":    "MFI scored specifically for continuation potential (not just high flow). "
                   "1.0 = MFI 55–75: buying pressure without overbought risk — ideal for continuation. "
                   "0.7 = MFI 75–80: strong inflow but approaching caution zone. "
                   "0.3 = MFI 80–90: overbought — reversal risk elevated. "
                   "0.0 = MFI >90 or <50: extreme overbought or no buying pressure.",
    },
    "NoBearDiv": {
        "label":   "No Bearish Divergence",
        "weight":  2,
        "desc":    "Checks that price highs and MFI highs are aligned (no bearish divergence). "
                   "1.0 = price and MFI both making higher highs — momentum confirmed. "
                   "0.5 = no clear divergence signal either way. "
                   "0.0 = bearish divergence detected: price made new high but MFI did not "
                   "(rally losing conviction — often precedes a reversal). "
                   "Measured over the last 20 trading days.",
    },
    "MA50Proximity": {
        "label":   "MA50 Proximity",
        "weight":  1,
        "desc":    "How close price is to the 50-day MA — rewards stocks near but above it. "
                   "1.0 = price 0–5% above MA50 (just reclaimed or sitting on support — low risk entry). "
                   "0.7 = price 5–10% above MA50 (healthy uptrend, not extended). "
                   "0.3 = price 10–20% above MA50 (extended — higher entry risk). "
                   "0.0 = price >20% above MA50 (very extended) or below MA50.",
    },
}

ALL_SECTORS = [
    "All Sectors",
    "Communication Services",
    "Consumer Discretionary",
    "Consumer Staples",
    "Energy",
    "Financials",
    "Health Care",
    "Industrials",
    "Information Technology",
    "Materials",
    "Real Estate",
    "Utilities",
]

EXCHANGES = {
    "S&P 500":  "sp500",
    "NYSE":     "nyse",
    "NASDAQ":   "nasdaq",
}

# Keywords used to detect and exclude non-company securities (ETFs, funds, trusts, etc.)
# These are matched against the security's longName / shortName from yfinance.
# ETF/fund keyword list — matched against longName/shortName (lowercase substring).
# IMPORTANT: only include terms that CANNOT appear in legitimate company names.
# Removed: "income", "trust", "preferred", "notes", "bond", "treasury", "index",
#          "fund", "portfolio", "commodity" — these match real operating companies.
ETF_KEYWORDS = [
    "etf", "ishares", "invesco", "vanguard", "spdr", "proshares",
    "direxion", "wisdomtree", "vaneck", "schwab select",
    "fidelity select", "global x", "ark ", "pimco",
    "debenture", "warrant",
]

def check_yfinance_health() -> dict:
    """
    Make a lightweight test request to yfinance to classify the current
    connection status. Returns a dict:
        status  : "fast" | "slow" | "rate_limited" | "timeout" | "blocked" | "unknown"
        latency : float (seconds) or None
        message : human-readable description
        fix     : short action string
    """
    import time as _time
    result = {"status": "unknown", "latency": None, "message": "", "fix": ""}
    try:
        t0   = _time.time()
        info = yf.Ticker("AAPL").fast_info   # lightest possible yfinance call
        # fast_info is a property — access one field to force the request
        _ = info.last_price
        latency = round(_time.time() - t0, 2)
        result["latency"] = latency

        if _ is None:
            result.update({
                "status":  "rate_limited",
                "message": "yfinance returned empty data — likely rate limited.",
                "fix":     "wait",
            })
        elif latency < 1.5:
            result.update({
                "status":  "fast",
                "message": f"Connection healthy ({latency}s response time).",
                "fix":     "none",
            })
        elif latency < 5.0:
            result.update({
                "status":  "slow",
                "message": f"Connection slow ({latency}s response time). "
                           "Reduce parallel workers to avoid overloading yfinance.",
                "fix":     "reduce_workers",
            })
        else:
            result.update({
                "status":  "slow",
                "message": f"Very slow response ({latency}s). High server load or network issue.",
                "fix":     "reduce_workers",
            })

    except Exception as e:
        err = str(e).lower()
        latency = round(_time.time() - t0, 2) if 't0' in dir() else None
        result["latency"] = latency
        if "timeout" in err or "timed out" in err or "read timed out" in err:
            result.update({
                "status":  "timeout",
                "message": "Connection timed out — yfinance server not responding.",
                "fix":     "retry",
            })
        elif "429" in err or "rate" in err or "too many" in err:
            result.update({
                "status":  "rate_limited",
                "message": "Rate limit hit (HTTP 429) — too many requests sent too fast.",
                "fix":     "wait",
            })
        elif "connection" in err or "network" in err or "resolve" in err or "refused" in err:
            result.update({
                "status":  "blocked",
                "message": "Network connection failed — Streamlit Cloud may be blocking yfinance.",
                "fix":     "sp500",
            })
        else:
            result.update({
                "status":  "unknown",
                "message": f"Unexpected error: {str(e)[:80]}",
                "fix":     "retry",
            })
    return result



# ───────────────────────────────────────────────────────────────
# CACHE HELPERS — must be defined before sidebar renders
# ───────────────────────────────────────────────────────────────
import pickle as _pickle

def _cache_size_str() -> str:
    total = 0
    for key in ("screener_cache", "analyze_fin_cache"):
        obj = st.session_state.get(key)
        if obj is None:
            continue
        try:
            total += len(_pickle.dumps(obj, protocol=2))
        except Exception:
            total += sys.getsizeof(obj)
    if total == 0:        return "0 KB"
    if total < 1024:      return f"{total} B"
    if total < 1024**2:   return f"{total/1024:.1f} KB"
    return f"{total/1024**2:.2f} MB"

def _clear_all_cache():
    for key in ("screener_cache", "analyze_fin_cache"):
        st.session_state.pop(key, None)

# ───────────────────────────────────────────────────────────────
# SESSION STATE DEFAULTS — set once on first load
# ───────────────────────────────────────────────────────────────
_defaults = {
    "slider_max_price":   250,
    "slider_min_score":   2.0,
    "tog_ma50":           "below",  # "off" | "below" | "above"
    "tog_range":          False,
    "slider_range_days":  30,
    "slider_range_pct":   10.0,
    "slider_mfi_period":  14,
    "analyze_history":    [],   # list of {ticker, name} dicts — most recent first
    "max_workers_val":    5,   # safe default; presets set to 19
    "tog_pe_filter":      False,
    "slider_pe_range":    (0, 50),   # tuple: (min_pe, max_pe)
    "tog_rev_filter":     False,
    "slider_rev_min":     0,
}
# Metric toggle/weight defaults
for _k, _cfg in METRICS.items():
    _defaults[f"tog_{_k}"] = True
    _defaults[f"wt_{_k}"]  = float(_cfg["weight"])
# RangePosScore off by default — only meaningful when range filter is on
_defaults["tog_RangePosScore"] = False
# New technical metrics — off by default so existing users aren't disrupted
for _k in ["RSI", "MACD", "GoldenCross", "MFISweetSpot", "NoBearDiv", "MA50Proximity"]:
    _defaults[f"tog_{_k}"] = False
    _defaults[f"wt_{_k}"]  = float(METRICS[_k]["weight"])

for _key, _val in _defaults.items():
    if _key not in st.session_state:
        st.session_state[_key] = _val

# ───────────────────────────────────────────────────────────────
# SIDEBAR — FILTERS  (Screener tab only)
# ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Screener Filters")

    # ── Magic Stock preset button ────────────────────────────────
    if st.button("✨ Magic Stock", use_container_width=True,
                 help="Finds stocks with the full continuation signal stack: "
                      "MFI in sweet spot (55–75), rising OBV, MACD accelerating, "
                      "golden cross confirmed, no bearish divergence, "
                      "price near MA50 support. The ideal 'going higher' setup."):
        # ── Filter settings ──────────────────────────────────────
        st.session_state["slider_max_price"]   = 500    # Don't cap price — quality stocks can be expensive
        st.session_state["slider_min_score"]   = 0.0    # Let the metric weights do the filtering
        st.session_state["tog_ma50"]           = "above"  # Must be in uptrend — above 50MA
        st.session_state["tog_range"]          = False  # Range filter off — we want trending stocks
        st.session_state["slider_range_days"]  = 20
        st.session_state["slider_range_pct"]   = 15.0
        st.session_state["slider_mfi_period"]  = 14
        # Turn off valuation filters — momentum setup doesn't require value conditions
        st.session_state["tog_pe_filter"]      = False
        st.session_state["tog_rev_filter"]     = False
        st.session_state["max_workers_val"]    = 8

        # ── Turn off all metrics first ────────────────────────────
        for _k in METRICS:
            st.session_state[f"tog_{_k}"] = False
            st.session_state[f"wt_{_k}"]  = 0.0

        # ── The full "going higher" signal stack ──────────────────
        # GoldenCross ×4 — 50MA above 200MA = institutional uptrend confirmed
        st.session_state["tog_GoldenCross"] = True
        st.session_state["wt_GoldenCross"]  = 4.0

        # MFISweetSpot ×4 — MFI 55–75: buying pressure without overbought risk
        st.session_state["tog_MFISweetSpot"] = True
        st.session_state["wt_MFISweetSpot"]  = 4.0

        # NoBearDiv ×3 — price and MFI both making higher highs = momentum confirmed
        st.session_state["tog_NoBearDiv"] = True
        st.session_state["wt_NoBearDiv"]  = 3.0

        # OBV ×3 — rising OBV = institutions accumulating, not distributing
        st.session_state["tog_OBV"]  = True
        st.session_state["wt_OBV"]   = 3.0

        # MACD ×3 — histogram positive and accelerating = short-term momentum confirmed
        st.session_state["tog_MACD"] = True
        st.session_state["wt_MACD"]  = 3.0

        # RSI ×2 — RSI 55–70: uptrend confirmed, room to run before overbought
        st.session_state["tog_RSI"] = True
        st.session_state["wt_RSI"]  = 2.0

        # MA50Proximity ×2 — price near MA50 support = low-risk entry, not extended
        st.session_state["tog_MA50Proximity"] = True
        st.session_state["wt_MA50Proximity"]  = 2.0

        # PCV ×2 — dominant up-day volume = buyers in control
        st.session_state["tog_PCV"]  = True
        st.session_state["wt_PCV"]   = 2.0

        # ROIC ×1 — quality filter: only fundamentally sound businesses
        st.session_state["tog_ROIC"] = True
        st.session_state["wt_ROIC"]  = 1.0

        st.rerun()

    # ── Insider Buying preset button ─────────────────────────────
    if st.button("🕵️ Insider Buying", use_container_width=True,
                 help="Searches for stocks showing signs of institutional/insider accumulation: "
                      "rising OBV while price is flat, strong money flow, dominant up-day volume. "
                      "Uses a 60-day range window — the sweet spot for catching slow institutional accumulation."): 
        # ── Filter settings ──────────────────────────────────────
        st.session_state["slider_max_price"]   = 200    # Wider price range — insiders buy mid-cap too
        st.session_state["slider_min_score"]   = 0.0
        st.session_state["tog_ma50"]           = "above"  # Insiders buying into strength — above 50MA
        st.session_state["tog_range"]          = True   # Insiders accumulate in a quiet range
        st.session_state["slider_range_days"]  = 60     # 60-day window — institutional accumulation takes months
        st.session_state["slider_range_pct"]   = 18.0   # Slightly wider — 60-day ranges naturally have more width
        st.session_state["slider_mfi_period"]  = 14     # Standard MFI period
        st.session_state["max_workers_val"]    = 8

        # ── Turn off all metrics first ────────────────────────────
        for _k in METRICS:
            st.session_state[f"tog_{_k}"] = False
            st.session_state[f"wt_{_k}"]  = 0.0

        # ── Insider accumulation metric weights ───────────────────
        # OBV ×4 — CORE: insiders can't hide volume. Rising OBV with flat price = pure accumulation
        st.session_state["tog_OBV"]  = True
        st.session_state["wt_OBV"]   = 4.0

        # GoldenCross ×3 — insiders buy into confirmed long-term uptrends (50MA > 200MA)
        st.session_state["tog_GoldenCross"] = True
        st.session_state["wt_GoldenCross"]  = 3.0

        # MFISweetSpot ×3 — sustained buying pressure without overbought risk
        st.session_state["tog_MFISweetSpot"] = True
        st.session_state["wt_MFISweetSpot"]  = 3.0

        # PCV ×3 — heavy up-day volume = insiders buying aggressively on certain days
        st.session_state["tog_PCV"]  = True
        st.session_state["wt_PCV"]   = 3.0

        # NoBearDiv ×2 — price and MFI aligned higher = accumulation is genuine
        st.session_state["tog_NoBearDiv"] = True
        st.session_state["wt_NoBearDiv"]  = 2.0

        # MACD ×2 — momentum building confirms the accumulation is starting to move price
        st.session_state["tog_MACD"] = True
        st.session_state["wt_MACD"]  = 2.0

        # RangePosScore ×2 — insiders buy near the low, not after it's already moved
        st.session_state["tog_RangePosScore"] = True
        st.session_state["wt_RangePosScore"]  = 2.0

        # ROIC ×2 — insiders buy companies they know are earning well on capital
        st.session_state["tog_ROIC"] = True
        st.session_state["wt_ROIC"]  = 2.0

        # RSI ×1 — confirm momentum is building without being overbought
        st.session_state["tog_RSI"] = True
        st.session_state["wt_RSI"]  = 1.0

        # EarningsGrowth ×1 — insiders know earnings direction ahead of the market
        st.session_state["tog_EarningsGrowth"] = True
        st.session_state["wt_EarningsGrowth"]  = 1.0

        st.rerun()

    st.divider()

    exchange = st.selectbox(
        "Exchange / Universe",
        options=list(EXCHANGES.keys()),
        index=0,
        help="S&P 500 = ~500 stocks (fast). NYSE/NASDAQ = 2,000–3,500 stocks (slow, 30–60+ min)."
    )

    sector = st.selectbox(
        "Market Sector",
        options=ALL_SECTORS,
        index=0,
        help="Filter to a specific GICS sector. Applied after scan using yfinance data."
    )

    max_price = st.slider(
        "Max Share Price ($)",
        min_value=10, max_value=1000, step=10,
        help="Exclude stocks above this price.",
        key="slider_max_price",
    )

    min_score = st.slider(
        "Min Hybrid Score",
        min_value=0.0, max_value=20.0, step=0.5,
        help="Only show stocks with a score above this threshold.",
        key="slider_min_score",
    )

    top_n = st.slider(
        "Max Results",
        min_value=5, max_value=100, value=30, step=5,
        help="Maximum number of stocks to display."
    )

    st.markdown("**50-Day Moving Average Filter**")
    ma50_mode = st.radio(
        "50-Day MA Filter",
        options=["off", "below", "above"],
        format_func=lambda x: {
            "off":   "⬜ No MA50 filter",
            "below": "🟢 Below 50-day MA  (pullbacks)",
            "above": "🔵 Above 50-day MA  (uptrends)",
        }[x],
        key="tog_ma50",
        label_visibility="collapsed",
        help="Filter stocks by their position relative to the 50-day moving average. "
             "'Below' finds stocks in pullbacks/consolidation. "
             "'Above' finds stocks already in an uptrend.",
    )
    use_ma50_filter = ma50_mode

    st.divider()
    st.header("📦 Price Range Filter")
    st.caption("Find stocks locked in a tight trading range — coiling before a breakout.")

    use_range_filter = st.toggle(
        "Enable Price Range Filter",
        help="When on, only shows stocks trading within a tight range over the selected period.",
        key="tog_range",
    )

    range_days = st.slider(
        "Range Lookback (days)",
        min_value=5, max_value=180, step=5,
        help="Number of trading days to measure the high/low range over.",
        key="slider_range_days",
    )

    range_max_pct = st.slider(
        "Max Range Width (%)",
        min_value=1.0, max_value=30.0, step=0.5,
        help="Maximum allowed spread between high and low as % of price. "
             "Lower = tighter range. 5–10% finds stocks in consolidation.",
        disabled=not use_range_filter,
        key="slider_range_pct",
    )

    st.divider()
    st.header("📐 Valuation & Growth Filters")

    use_pe_filter = st.toggle(
        "Enable P/E Ratio Filter",
        help="Only show stocks whose trailing P/E falls within the selected range. "
             "Stocks with no P/E (negative earnings) are excluded when this is on.",
        key="tog_pe_filter",
    )
    pe_min, pe_max = st.slider(
        "P/E Ratio Range",
        min_value=0, max_value=200, step=1,
        disabled=not use_pe_filter,
        help="Filter to stocks with a trailing P/E between these two values. "
             "0–15 = value territory · 15–25 = fair value · 25–50 = growth premium · 50+ = high growth/speculative.",
        key="slider_pe_range",
    )

    st.divider()
    # TTM revenue growth filter in 15% increments
    REV_GROWTH_STEPS = [0, 15, 30, 45, 60, 75, 90]
    use_rev_filter = st.toggle(
        "Enable TTM Revenue Growth Filter",
        help="Only show stocks whose trailing-12-month revenue growth meets the minimum threshold.",
        key="tog_rev_filter",
    )
    rev_min_idx = st.select_slider(
        "Min TTM Revenue Growth",
        options=REV_GROWTH_STEPS,
        format_func=lambda x: f"{x}%+",
        disabled=not use_rev_filter,
        help="Minimum trailing-12-month revenue growth rate. Steps are in 15% increments.",
        key="slider_rev_min",
    )

    st.divider()
    st.header("📊 Metrics")
    st.caption("Toggle on/off and adjust weight in the final score.")

    metric_enabled = {}
    metric_weight  = {}

    # ── Fundamental metrics ──────────────────────────────────────
    st.markdown("**Fundamental**")
    for key in ["OE_Yield", "ROIC", "ROIC_Trend", "RevenueGrowth", "EarningsGrowth", "Piotroski"]:
        cfg = METRICS[key]
        col_a, col_b = st.columns([1, 2])
        with col_a:
            metric_enabled[key] = st.toggle(cfg["label"], key=f"tog_{key}")
        with col_b:
            metric_weight[key] = st.slider(
                "Weight", min_value=0.0, max_value=5.0, step=0.5,
                key=f"wt_{key}",
                disabled=not metric_enabled[key],
                label_visibility="collapsed",
            )
        st.caption(cfg["desc"])

    # ── Volume / buying pressure metrics ────────────────────────
    st.markdown("**Volume & Buying Pressure**")
    for key in ["OBV", "MFI", "PCV"]:
        cfg = METRICS[key]
        col_a, col_b = st.columns([1, 2])
        with col_a:
            metric_enabled[key] = st.toggle(cfg["label"], key=f"tog_{key}")
        with col_b:
            metric_weight[key] = st.slider(
                "Weight", min_value=0.0, max_value=5.0, step=0.5,
                key=f"wt_{key}",
                disabled=not metric_enabled[key],
                label_visibility="collapsed",
            )
        st.caption(cfg["desc"])

    # ── Technical / momentum metrics ─────────────────────────
    st.markdown("**Technical & Momentum**")
    for key in ["RSI", "MACD", "GoldenCross", "MFISweetSpot", "NoBearDiv", "MA50Proximity"]:
        cfg = METRICS[key]
        col_a, col_b = st.columns([1, 2])
        with col_a:
            metric_enabled[key] = st.toggle(cfg["label"], key=f"tog_{key}")
        with col_b:
            metric_weight[key] = st.slider(
                "Weight", min_value=0.0, max_value=5.0, step=0.5,
                key=f"wt_{key}",
                disabled=not metric_enabled[key],
                label_visibility="collapsed",
            )
        st.caption(cfg["desc"])

    # ── Range position metric ─────────────────────────────────
    st.markdown("**Range & Breakout Setup**")
    for key in ["RangePosScore"]:
        cfg = METRICS[key]
        col_a, col_b = st.columns([1, 2])
        with col_a:
            metric_enabled[key] = st.toggle(cfg["label"], key=f"tog_{key}")
        with col_b:
            metric_weight[key] = st.slider(
                "Weight", min_value=0.0, max_value=5.0, step=0.5,
                key=f"wt_{key}",
                disabled=not metric_enabled[key],
                label_visibility="collapsed",
            )
        st.caption(cfg["desc"])

    st.divider()
    st.header("🔧 Performance")
    max_workers = st.slider(
        "Parallel Workers",
        min_value=1, max_value=20, step=1,
        key="max_workers_val",
        help="Lower values reduce rate-limiting errors on cloud. Recommended: 3–5 on Streamlit Cloud, 10+ on local PC."
    )

    mfi_period = st.slider(
        "MFI Period (days)",
        min_value=7, max_value=30, step=1,
        help="Money Flow Index lookback window.",
        key="slider_mfi_period",
    )

    st.divider()
    st.markdown("**📡 yfinance Connection**")

    # ── Live health check ─────────────────────────────────────────
    _do_check = st.button("🔍 Check Connection", use_container_width=True,
                          help="Test the yfinance data connection before scanning.")
    if _do_check or st.session_state.get("_yf_health"):
        if _do_check:
            with st.spinner("Testing connection..."):
                _health = check_yfinance_health()
            st.session_state["_yf_health"] = _health
        else:
            _health = st.session_state["_yf_health"]

        _st  = _health["status"]
        _lat = _health["latency"]
        _msg = _health["message"]
        _fix = _health["fix"]

        # ── Status indicator ──────────────────────────────────────
        _STATUS_UI = {
            "fast":         ("🟢", "Healthy",      "success"),
            "slow":         ("🟡", "Slow",          "warning"),
            "rate_limited": ("🔴", "Rate Limited",  "error"),
            "timeout":      ("🔴", "Timed Out",     "error"),
            "blocked":      ("⛔", "Blocked",       "error"),
            "unknown":      ("⚪", "Unknown",       "warning"),
        }
        _icon, _label, _sev = _STATUS_UI.get(_st, ("⚪", _st, "warning"))
        _lat_str = f" · {_lat}s" if _lat is not None else ""

        if _sev == "success":
            st.success(f"{_icon} **{_label}**{_lat_str} — {_msg}")
        elif _sev == "warning":
            st.warning(f"{_icon} **{_label}**{_lat_str} — {_msg}")
        else:
            st.error(f"{_icon} **{_label}**{_lat_str} — {_msg}")

        # ── Fix buttons ───────────────────────────────────────────
        if _fix == "wait":
            st.caption("⏳ **Fix:** Wait 30–60 seconds for the rate limit to reset, then scan again.")
            if st.button("⏱️ Wait 30s then Re-check", use_container_width=True,
                         key="fix_wait_btn"):
                import time as _t; _t.sleep(30)
                _health = check_yfinance_health()
                st.session_state["_yf_health"] = _health
                st.rerun()
            if st.button("⬇️ Reduce Workers to 3", use_container_width=True,
                         key="fix_reduce_btn"):
                st.session_state["max_workers_val"] = 3
                st.session_state.pop("_yf_health", None)
                st.rerun()

        elif _fix == "reduce_workers":
            st.caption("⚡ **Fix:** Reduce parallel workers to lower request pressure.")
            if st.button("⬇️ Set Workers to 3", use_container_width=True,
                         key="fix_slow_btn"):
                st.session_state["max_workers_val"] = 3
                st.session_state.pop("_yf_health", None)
                st.rerun()

        elif _fix == "timeout":
            st.caption("🔄 **Fix:** Retry — transient timeout, server may be temporarily overloaded.")
            if st.button("🔄 Re-check Connection", use_container_width=True,
                         key="fix_timeout_btn"):
                st.session_state.pop("_yf_health", None)
                st.rerun()

        elif _fix == "sp500":
            st.caption("🔄 **Fix:** Switch to S&P 500 (smaller universe), or wait for network to recover.")
            if st.button("📊 Switch to S&P 500", use_container_width=True,
                         key="fix_sp500_btn"):
                st.session_state.pop("_yf_health", None)
                st.rerun()

        if st.button("✕ Clear Status", use_container_width=True, key="clear_health_btn"):
            st.session_state.pop("_yf_health", None)
            st.rerun()

    st.divider()
    run_btn = st.button("🚀 Run Screener", use_container_width=True, type="primary")

    # ── Cache status panel ────────────────────────────────────────
    st.divider()
    _sc  = st.session_state.get("screener_cache", {})
    _fca = st.session_state.get("analyze_fin_cache", {})
    _sz  = _cache_size_str()

    if _sc.get("results") or _fca:
        _exch_name    = next((k for k, v in EXCHANGES.items() if v == _sc.get("exchange_key")), "?")
        _scan_tickers = len(_sc.get("results", []))
        _fin_tickers  = len(_fca)
        # Count how many rows have raw history stored
        _with_hist    = sum(1 for r in _sc.get("results", []) if r.get("_hist"))

        st.markdown("**💾 Cache Status**")
        ci1, ci2 = st.columns(2)
        ci1.metric("Screener Tickers", _scan_tickers if _scan_tickers else "—")
        ci2.metric("Analyzed Stocks",  _fin_tickers  if _fin_tickers  else "—")
        ci3, ci4 = st.columns(2)
        ci3.metric("Cache Size",   _sz)
        ci4.metric("Exchange",     _exch_name if _scan_tickers else "—")
        if _with_hist:
            st.caption(f"✅ {_with_hist}/{_scan_tickers} tickers have raw history — "
                       f"MFI period & range window recompute instantly with no re-download.")

        if _sc.get("scanned_at"):
            st.caption(f"🕐 Last scan: {_sc['scanned_at']}  ·  "
                       f"Cache clears automatically when you close the browser tab.")

        if st.button("🗑️ Clear All Cache", use_container_width=True, key="clear_all_cache_btn"):
            _clear_all_cache()
            st.success("All cache cleared.")
            st.rerun()
    else:
        st.markdown("**💾 Cache Status**")
        st.caption("No cache yet — run a scan or analyze a stock to populate it.  "
                   "Cache is stored in your browser session and clears automatically when the tab is closed.")

# ───────────────────────────────────────────────────────────────

def get_volume_signals(hist_df, mfi_period):
    """
    Compute OBV / MFI / PCV from a pre-fetched OHLCV DataFrame.

    OBV  — On-Balance Volume trend over last 20 days (1.0 = rising, 0.0 = falling)
    MFI  — Money Flow Index scaled above 50 neutral (0.0–1.0)
    PCV  — Price-Confirmed Volume fraction on up-days, scaled above 50% baseline (0.0–1.0)

    Accepts either a stock object (legacy) or a DataFrame directly.
    """
    default = {"OBV": 0.0, "MFI": 0.0, "PCV": 0.0}
    try:
        # Accept either a raw DataFrame or a yfinance Ticker object (fallback)
        if isinstance(hist_df, pd.DataFrame):
            hist = hist_df.copy()
        else:
            # Legacy path — stock object passed directly
            hist = hist_df.history(period="1y")
            if hist.empty or len(hist) < mfi_period + 5:
                hist = hist_df.history(period="3mo")
            if hist.empty or len(hist) < mfi_period + 5:
                hist = hist_df.history(period="1mo")

        if hist.empty or len(hist) < 10:
            return default

        close, high, low, vol = hist["Close"], hist["High"], hist["Low"], hist["Volume"]

        # Drop any rows where volume is 0 (halted/illiquid bars corrupt MFI)
        mask = vol > 0
        close, high, low, vol = close[mask], high[mask], low[mask], vol[mask]
        if len(close) < 10:
            return default

        # ── OBV ──────────────────────────────────────────────────
        direction = np.sign(close.diff().fillna(0))
        obv       = (direction * vol).cumsum()
        obv_window = min(20, len(obv))
        obv_slope = np.polyfit(range(obv_window), obv.iloc[-obv_window:].values, 1)[0]
        obv_score = 1.0 if obv_slope > 0 else 0.0

        # ── MFI ──────────────────────────────────────────────────
        # Use the smaller of mfi_period or half the available bars to handle short histories
        effective_period = min(mfi_period, max(5, len(close) // 2))
        typical_price = (high + low + close) / 3
        raw_mf  = typical_price * vol
        tp_diff = typical_price.diff()

        pos_mf  = raw_mf.where(tp_diff > 0, 0).rolling(effective_period).sum()
        neg_mf  = raw_mf.where(tp_diff < 0, 0).rolling(effective_period).sum()

        # Replace 0 neg_mf with NaN to avoid divide-by-zero
        # When neg_mf is 0, all money flow is positive → MFI = 100 → score = 1.0
        all_positive = neg_mf == 0
        mfr = pos_mf / neg_mf.replace(0, np.nan)
        mfi_series = 100 - (100 / (1 + mfr))
        # Where neg_mf was 0 and pos_mf > 0, MFI should be 100
        mfi_series = mfi_series.where(~(all_positive & (pos_mf > 0)), 100.0)

        # Take last valid value — walk back up to 5 bars if last is NaN
        mfi_val = None
        for _i in range(1, 6):
            candidate = mfi_series.iloc[-_i]
            if pd.notnull(candidate):
                mfi_val = float(candidate)
                break

        mfi_score = round(mfi_val / 100.0, 4) if mfi_val is not None else 0.0

        # ── PCV ──────────────────────────────────────────────────
        pcv_window = min(20, len(close))
        recent     = pd.DataFrame({"Close": close, "Volume": vol}).iloc[-pcv_window:]
        recent_up  = recent["Close"] > recent["Close"].shift(1)
        up_vol     = recent.loc[recent_up, "Volume"].sum()
        total_vol  = recent["Volume"].sum()
        pcv_ratio  = up_vol / total_vol if total_vol > 0 else 0.5
        pcv_score  = max(0.0, (pcv_ratio - 0.5) / 0.5)

        return {
            "OBV": round(obv_score, 4),
            "MFI": round(mfi_score, 4),
            "PCV": round(pcv_score, 4),
        }
    except:
        return default


def calculate_technical_signals(hist: pd.DataFrame) -> dict:
    """
    Compute RSI, MACD, Golden Cross, MFI Sweet Spot, No-Bearish-Divergence,
    and MA50 Proximity from a pre-fetched OHLCV DataFrame.
    All scores are 0.0–1.0. Returns zeros on insufficient data.
    """
    default = {
        "RSI":           0.0,
        "MACD":          0.0,
        "GoldenCross":   0.0,
        "MFISweetSpot":  0.0,
        "NoBearDiv":     0.5,
        "MA50Proximity": 0.0,
    }
    try:
        if hist.empty or len(hist) < 26:
            return default

        close = hist["Close"].dropna()
        if len(close) < 26:
            return default

        # ── RSI (14-period) ──────────────────────────────────────
        rsi_score = 0.0
        try:
            delta = close.diff()
            gain  = delta.clip(lower=0).rolling(14).mean()
            loss  = (-delta.clip(upper=0)).rolling(14).mean()
            rs    = gain / loss.replace(0, np.nan)
            rsi_s = 100 - (100 / (1 + rs))
            rsi_val = rsi_s.dropna().iloc[-1] if not rsi_s.dropna().empty else None
            if rsi_val is not None:
                if   55 <= rsi_val <= 70: rsi_score = 1.0   # sweet spot
                elif 50 <= rsi_val <  55: rsi_score = 0.6   # building
                elif 70 <  rsi_val <= 80: rsi_score = 0.2   # extended
                else:                     rsi_score = 0.0   # overbought or no trend
        except Exception:
            pass

        # ── MACD (12/26/9) ───────────────────────────────────────
        macd_score = 0.0
        try:
            ema12   = close.ewm(span=12, adjust=False).mean()
            ema26   = close.ewm(span=26, adjust=False).mean()
            macd_ln = ema12 - ema26
            signal  = macd_ln.ewm(span=9, adjust=False).mean()
            hist_m  = macd_ln - signal
            if len(hist_m.dropna()) >= 2:
                h_now  = hist_m.dropna().iloc[-1]
                h_prev = hist_m.dropna().iloc[-2]
                if h_now > 0 and h_now > h_prev:   macd_score = 1.0  # positive & accelerating
                elif h_now > 0:                     macd_score = 0.6  # positive but decelerating
                elif h_now > h_prev:                macd_score = 0.2  # negative but improving
                else:                               macd_score = 0.0
        except Exception:
            pass

        # ── Golden Cross (50MA vs 200MA) ─────────────────────────
        golden_score = 0.0
        try:
            if len(close) >= 200:
                ma50  = close.rolling(50).mean().iloc[-1]
                ma200 = close.rolling(200).mean().iloc[-1]
                if pd.notnull(ma50) and pd.notnull(ma200) and ma200 > 0:
                    diff_pct = (ma50 - ma200) / ma200
                    if   diff_pct >  0.02: golden_score = 1.0   # golden cross confirmed
                    elif diff_pct >= -0.02: golden_score = 0.5  # within 2% — imminent
                    else:                  golden_score = 0.0   # death cross
        except Exception:
            pass

        # ── MFI Sweet Spot ───────────────────────────────────────
        mfi_sweet = 0.0
        try:
            high_s = hist["High"].dropna()
            low_s  = hist["Low"].dropna()
            vol_s  = hist["Volume"].dropna()
            # Align all series
            idx = close.index.intersection(high_s.index).intersection(low_s.index).intersection(vol_s.index)
            c2, h2, l2, v2 = close[idx], high_s[idx], low_s[idx], vol_s[idx]
            v2 = v2.where(v2 > 0)
            tp  = (h2 + l2 + c2) / 3
            rmf = tp * v2
            tp_diff = tp.diff()
            period = 14
            pos_mf = rmf.where(tp_diff > 0, 0).rolling(period).sum()
            neg_mf = rmf.where(tp_diff < 0, 0).rolling(period).sum()
            mfr2   = pos_mf / neg_mf.replace(0, np.nan)
            mfi2   = 100 - (100 / (1 + mfr2))
            mfi_v  = mfi2.dropna().iloc[-1] if not mfi2.dropna().empty else None
            if mfi_v is not None:
                if   55 <= mfi_v <= 75: mfi_sweet = 1.0
                elif 75 <  mfi_v <= 80: mfi_sweet = 0.7
                elif 80 <  mfi_v <= 90: mfi_sweet = 0.3
                else:                   mfi_sweet = 0.0
        except Exception:
            pass

        # ── No Bearish Divergence (20-day window) ────────────────
        no_bear_div = 0.5  # neutral default
        try:
            window = 20
            if len(close) >= window * 2:
                # Split recent history into two halves
                mid   = len(close) - window
                c_old = close.iloc[mid - window : mid]
                c_new = close.iloc[mid:]
                price_higher = c_new.max() > c_old.max()

                # Recompute MFI for divergence check
                high_s = hist["High"]
                low_s  = hist["Low"]
                vol_s  = hist["Volume"]
                tp_d   = (high_s + low_s + close) / 3
                rmf_d  = (tp_d * vol_s).where(vol_s > 0)
                tpd_d  = tp_d.diff()
                pos_d  = rmf_d.where(tpd_d > 0, 0).rolling(14).sum()
                neg_d  = rmf_d.where(tpd_d < 0, 0).rolling(14).sum()
                mfr_d  = pos_d / neg_d.replace(0, np.nan)
                mfi_d  = (100 - (100 / (1 + mfr_d))).dropna()
                if len(mfi_d) >= window * 2:
                    mfi_old = mfi_d.iloc[-(window * 2):-window]
                    mfi_new = mfi_d.iloc[-window:]
                    mfi_higher = mfi_new.max() > mfi_old.max()
                    if price_higher and mfi_higher:     no_bear_div = 1.0  # both higher — confirmed
                    elif price_higher and not mfi_higher: no_bear_div = 0.0  # bearish divergence
                    else:                               no_bear_div = 0.5
        except Exception:
            pass

        # ── MA50 Proximity ────────────────────────────────────────
        ma50_prox = 0.0
        try:
            if len(close) >= 50:
                ma50_v  = close.rolling(50).mean().iloc[-1]
                price_v = close.iloc[-1]
                if pd.notnull(ma50_v) and ma50_v > 0:
                    pct_above = (price_v - ma50_v) / ma50_v
                    if   0.0  <= pct_above <= 0.05: ma50_prox = 1.0
                    elif 0.05 <  pct_above <= 0.10: ma50_prox = 0.7
                    elif 0.10 <  pct_above <= 0.20: ma50_prox = 0.3
                    else:                            ma50_prox = 0.0  # extended or below
        except Exception:
            pass

        return {
            "RSI":           round(rsi_score,    4),
            "MACD":          round(macd_score,   4),
            "GoldenCross":   round(golden_score, 4),
            "MFISweetSpot":  round(mfi_sweet,    4),
            "NoBearDiv":     round(no_bear_div,  4),
            "MA50Proximity": round(ma50_prox,    4),
        }
    except Exception:
        return default


def calculate_price_range(hist_df, range_days: int) -> dict:
    """
    Calculates the price range over the last `range_days` trading days.
    Accepts either a pre-fetched OHLCV DataFrame or a yfinance Ticker (legacy).

    Returns:
        RangeHigh  — highest closing price in the period
        RangeLow   — lowest closing price in the period
        RangePct   — range width as % of midpoint price  (High-Low) / Mid × 100
                     Lower = tighter range = stock is consolidating
        RangePos   — where today's price sits in the range  0.0 = at low, 1.0 = at high
                     0.5 = midpoint; useful for spotting stocks near support (low end)
    """
    default = {"RangeHigh": None, "RangeLow": None, "RangePct": None, "RangePos": None}
    try:
        if isinstance(hist_df, pd.DataFrame):
            hist = hist_df
        else:
            hist = hist_df.history(period="1y")
        if hist.empty or len(hist) < range_days:
            return default
        window   = hist["Close"].iloc[-range_days:]
        high     = round(window.max(), 2)
        low      = round(window.min(), 2)
        mid      = (high + low) / 2
        if mid == 0:
            return default
        rng_pct  = round((high - low) / mid * 100, 2)
        current  = window.iloc[-1]
        rng_pos  = round((current - low) / (high - low), 4) if (high - low) > 0 else 0.5
        return {
            "RangeHigh": high,
            "RangeLow":  low,
            "RangePct":  rng_pct,
            "RangePos":  rng_pos,
        }
    except:
        return default


def calculate_piotroski(fin, bal, cf):
    """Accepts pre-fetched financials/balance/cashflow DataFrames — no extra API calls."""
    try:
        score = 0
        roa = fin.loc["Net Income"] / bal.loc["Total Assets"]
        if roa.iloc[0] > 0: score += 1
        if cf.loc["Operating Cash Flow"].iloc[0] > 0: score += 1
        if roa.iloc[0] > roa.iloc[1]: score += 1
        if cf.loc["Operating Cash Flow"].iloc[0] > fin.loc["Net Income"].iloc[0]: score += 1
        if bal.loc["Long Term Debt"].iloc[0] < bal.loc["Long Term Debt"].iloc[1]: score += 1
        if (bal.loc["Current Assets"].iloc[0] / bal.loc["Current Liabilities"].iloc[0]) > \
           (bal.loc["Current Assets"].iloc[1] / bal.loc["Current Liabilities"].iloc[1]): score += 1
        return score
    except:
        return None


def get_owner_earnings(cf, fin, info):
    """Accepts pre-fetched cashflow/financials DataFrames — no extra API calls."""
    try:
        oe = fin.loc["Net Income"].iloc[0] + cf.loc["Depreciation"].iloc[0] - abs(cf.loc["Capital Expenditure"].iloc[0])
        mc = info.get("marketCap")
        return oe, (oe / mc if mc else None)
    except:
        return None, None


def _get_fin_value(fin, *labels):
    for label in labels:
        if label in fin.index:
            return fin.loc[label]
    return None

def _get_bal_value(bal, *labels):
    for label in labels:
        if label in bal.index:
            return bal.loc[label]
    return None

def calculate_roic(fin, bal):
    """Accepts pre-fetched financials/balance DataFrames — no extra API calls."""
    try:
        ebit_s = _get_fin_value(fin, "EBIT", "Ebit", "Operating Income", "OperatingIncome", "EBITDA", "Ebitda")
        if ebit_s is None: return None
        ebit = ebit_s.iloc[0]
        assets_s = _get_bal_value(bal, "Total Assets", "TotalAssets")
        if assets_s is None: return None
        liab_s = _get_bal_value(bal, "Total Current Liabilities", "TotalCurrentLiabilities", "Current Liabilities", "CurrentLiabilities")
        if liab_s is None: return None
        cash_s = _get_bal_value(bal, "Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments", "CashAndCashEquivalents", "Cash")
        cash = cash_s.iloc[0] if cash_s is not None else 0
        ic = assets_s.iloc[0] - liab_s.iloc[0] - cash
        if pd.isna(ic) or ic == 0: return None
        roic = ebit * 0.79 / ic
        return None if pd.isna(roic) else roic
    except:
        return None


def calculate_roic_trend(fin, bal):
    """Accepts pre-fetched financials/balance DataFrames — no extra API calls."""
    try:
        ebit_s = _get_fin_value(fin, "EBIT", "Ebit", "Operating Income", "OperatingIncome", "EBITDA", "Ebitda")
        if ebit_s is None or len(ebit_s) < 2: return None
        assets_s = _get_bal_value(bal, "Total Assets", "TotalAssets")
        liab_s   = _get_bal_value(bal, "Total Current Liabilities", "TotalCurrentLiabilities", "Current Liabilities", "CurrentLiabilities")
        if assets_s is None or liab_s is None: return None
        if len(assets_s) < 2 or len(liab_s) < 2: return None
        def roic_at(i):
            ic = assets_s.iloc[i] - liab_s.iloc[i]
            if pd.isna(ic) or ic == 0: return None
            v = ebit_s.iloc[i] / ic
            return None if pd.isna(v) else v
        r0, r1 = roic_at(0), roic_at(1)
        return (r0 - r1) if (r0 is not None and r1 is not None) else None
    except:
        return None


def process_ticker(args):
    """
    Fetch all data for one ticker in a single optimised pass.

    Optimisations vs naive approach
    --------------------------------
    • stock.info        — 1 call  (price, sector, fundamentals, revenue/earnings growth)
    • stock.history(2y) — 1 call  (OHLCV for ALL technical indicators)
    • stock.financials  — 1 call  (fetched once, passed to all 4 compute functions)
    • stock.balance_sheet — 1 call (shared across roic, roic_trend, piotroski)
    • stock.cashflow    — 1 call  (shared across owner_earnings, piotroski)
    Total: 5 HTTP requests per ticker  (down from up to 12 in naive version)

    Raw OHLCV stored in _hist so MFI/range can be recomputed from cache with
    any sidebar setting change — zero re-downloads needed.
    """
    t, mfi_period, range_days = args
    for attempt in range(3):
        try:
            if attempt > 0:
                # Exponential backoff with jitter — stagger retries across workers
                import random
                time.sleep(attempt * 2.0 + random.uniform(0, 1.0))

            stock = yf.Ticker(t)

            # ── 1. Info — price, sector, fundamentals ─────────────
            info = stock.info
            # yfinance returns a minimal dict (< 5 keys) when rate-limited or
            # when the ticker is invalid. Distinguish the two:
            # rate-limited dicts often have exactly 1-2 keys like {"trailingPegRatio": None}
            # invalid tickers return {} or {"regularMarketPrice": None}
            if not info or len(info) < 5:
                return None
            if is_etf_or_fund(info):
                return None
            price = info.get("currentPrice") or info.get("regularMarketPrice")
            if price is None:
                return None

            # ── 2. OHLCV history — single fetch, all indicators ───
            hist = stock.history(period="2y")
            if hist.empty:
                hist = stock.history(period="1y")
            if hist.empty:
                hist = stock.history(period="3mo")

            # ── 3. Financial statements — fetched ONCE each ───────
            try:
                fin = stock.financials
            except Exception:
                fin = pd.DataFrame()
            try:
                bal = stock.balance_sheet
            except Exception:
                bal = pd.DataFrame()
            try:
                cf = stock.cashflow
            except Exception:
                cf = pd.DataFrame()

            # ── 4. Compute all indicators — no additional API calls
            vol_signals    = get_volume_signals(hist, mfi_period)
            tech_signals   = calculate_technical_signals(hist)
            range_data     = calculate_price_range(hist, range_days)
            ma50           = round(hist["Close"].rolling(50).mean().iloc[-1], 2) if len(hist) >= 50 else None
            owner_earnings, oe_yield = get_owner_earnings(cf, fin, info)
            roic           = calculate_roic(fin, bal)
            roic_trend     = calculate_roic_trend(fin, bal)
            piotroski      = calculate_piotroski(fin, bal, cf)

            # ── 5. Compact OHLCV cache ────────────────────────────
            hist_cache = {
                "dates":  hist.index.strftime("%Y-%m-%d").tolist(),
                "open":   hist["Open"].tolist(),
                "high":   hist["High"].tolist(),
                "low":    hist["Low"].tolist(),
                "close":  hist["Close"].tolist(),
                "volume": hist["Volume"].tolist(),
            }

            return {
                "Ticker":         t,
                "Sector":         info.get("sector", "Unknown"),
                "Price":          price,
                "MarketCap":      info.get("marketCap"),
                "P/E":            info.get("trailingPE"),
                "OwnerEarnings":  owner_earnings,
                "OE_Yield":       oe_yield,
                "ROIC":           roic,
                "ROIC_Trend":     roic_trend,
                "RevenueGrowth":  info.get("revenueGrowth"),
                "EarningsGrowth": info.get("earningsGrowth"),
                "Piotroski":      piotroski,
                "MA50":           ma50,
                "OBV":            vol_signals["OBV"],
                "MFI":            vol_signals["MFI"],
                "PCV":            vol_signals["PCV"],
                "RSI":            tech_signals["RSI"],
                "MACD":           tech_signals["MACD"],
                "GoldenCross":    tech_signals["GoldenCross"],
                "MFISweetSpot":   tech_signals["MFISweetSpot"],
                "NoBearDiv":      tech_signals["NoBearDiv"],
                "MA50Proximity":  tech_signals["MA50Proximity"],
                "RangeHigh":      range_data["RangeHigh"],
                "RangeLow":       range_data["RangeLow"],
                "RangePct":       range_data["RangePct"],
                "RangePos":       range_data["RangePos"],
                "_hist":          hist_cache,
            }
        except Exception:
            if attempt == 2:
                return None
            continue
    return None


def _hist_from_cache(row: dict) -> pd.DataFrame:
    """Reconstruct a history DataFrame from the compact cache dict stored in a result row."""
    hc = row.get("_hist")
    if not hc:
        return pd.DataFrame()
    try:
        df = pd.DataFrame({
            "Open":   hc["open"],
            "High":   hc["high"],
            "Low":    hc["low"],
            "Close":  hc["close"],
            "Volume": hc["volume"],
        }, index=pd.to_datetime(hc["dates"]))
        return df
    except Exception:
        return pd.DataFrame()


def recompute_indicators(results: list, mfi_period: int, range_days: int) -> list:
    """
    Re-derive MFI, OBV, PCV, MA50, and range metrics from cached raw history
    using the current sidebar settings — zero network requests.
    Returns a new list of result dicts with updated indicator values.
    """
    updated = []
    for row in results:
        hist = _hist_from_cache(row)
        if hist.empty:
            updated.append(row)
            continue
        new_row = dict(row)
        vol  = get_volume_signals(hist, mfi_period)
        rng  = calculate_price_range(hist, range_days)
        ma50 = round(hist["Close"].rolling(50).mean().iloc[-1], 2) if len(hist) >= 50 else None
        tech = calculate_technical_signals(hist)
        new_row.update({
            "MA50":          ma50,
            "OBV":           vol["OBV"],
            "MFI":           vol["MFI"],
            "PCV":           vol["PCV"],
            "RSI":           tech["RSI"],
            "MACD":          tech["MACD"],
            "GoldenCross":   tech["GoldenCross"],
            "MFISweetSpot":  tech["MFISweetSpot"],
            "NoBearDiv":     tech["NoBearDiv"],
            "MA50Proximity": tech["MA50Proximity"],
            "RangeHigh":     rng["RangeHigh"],
            "RangeLow":      rng["RangeLow"],
            "RangePct":      rng["RangePct"],
            "RangePos":      rng["RangePos"],
        })
        updated.append(new_row)
    return updated

# ───────────────────────────────────────────────────────────────
# TICKER LOADERS
# ───────────────────────────────────────────────────────────────

def _fetch_exchange_tickers(exchange: str) -> list:
    """
    Fetch ALL tickers for NYSE or NASDAQ.

    Source 1 — NASDAQ Trader symbol directory (official, pipe-delimited, no pagination)
      nasdaqlisted.txt  ~3 300 NASDAQ tickers in one request
      otherlisted.txt   ~3 000 NYSE/AMEX/Arca tickers in one request
    Source 2 — NASDAQ screener API, paginated with explicit offset loop
    Source 3 — SEC EDGAR company_tickers_exchange.json
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; stockscreener/2.0)",
        "Accept":     "text/plain,application/json,*/*",
    }

    # ── Source 1: NASDAQ Trader symbol directory (correct domain) ──
    # www.nasdaqtrader.com/dynamic/SymDir/ — official NASDAQ data feeds
    # Single GET, no pagination, pipe-delimited, ETF + test-issue flags built in.
    try:
        if exchange == "nasdaq":
            url = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
        else:
            url = "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"

        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        lines = [l for l in resp.text.strip().splitlines() if l.strip()]

        if len(lines) > 2:
            header = [h.strip() for h in lines[0].split("|")]
            sym_i  = header.index("Symbol")     if "Symbol"     in header else 0
            etf_i  = header.index("ETF")        if "ETF"        in header else None
            test_i = header.index("Test Issue") if "Test Issue" in header else None
            exch_i = header.index("Exchange")   if "Exchange"   in header else None

            tickers = []
            for line in lines[1:]:
                # Last line is a file-creation date stamp — skip it
                if line.startswith("File Creation Time"):
                    continue
                parts = line.split("|")
                if len(parts) <= sym_i:
                    continue
                sym = parts[sym_i].strip()
                if not sym or sym == "Symbol":
                    continue
                if etf_i  is not None and len(parts) > etf_i  and parts[etf_i].strip()  == "Y":
                    continue
                if test_i is not None and len(parts) > test_i and parts[test_i].strip() == "Y":
                    continue
                # otherlisted: filter by exchange code
                # N = NYSE, A = NYSE American (AMEX), P = NYSE Arca, Z = BATS, V = Investors Exchange
                if exch_i is not None and exchange == "nyse":
                    exch_val = parts[exch_i].strip() if len(parts) > exch_i else ""
                    if exch_val not in ("N", "A", "P", "Z", "V"):
                        continue
                tickers.append(sym.replace(".", "-"))

            if len(tickers) >= 500:
                return tickers
    except Exception:
        pass

    # ── Source 2: NASDAQ screener API, fully paginated ────────────
    # Read totalrecords, then loop through all offset pages explicitly.
    # Do NOT break on empty page_rows — always loop to total.
    try:
        base = (f"https://api.nasdaq.com/api/screener/stocks"
                f"?tableonly=true&limit=1000&exchange={exchange}&offset=")
        all_rows = []

        resp  = requests.get(base + "0", headers=headers, timeout=30)
        resp.raise_for_status()
        table = resp.json().get("data", {}).get("table", {}) or {}

        # totalrecords arrives as a string, sometimes with commas
        raw_total = str(table.get("totalrecords") or "0")
        total     = int(raw_total.replace(",", "").strip() or "0")
        first_rows = table.get("rows") or []
        all_rows.extend(first_rows)

        # Fetch every remaining page — do NOT break early on empty response
        for offset in range(1000, total, 1000):
            try:
                pr = requests.get(base + str(offset), headers=headers, timeout=30)
                pr.raise_for_status()
                page_rows = (pr.json().get("data", {}).get("table", {}) or {}).get("rows") or []
                all_rows.extend(page_rows)
            except Exception:
                continue   # skip one bad page, keep going

        tickers = [row["symbol"].strip() for row in all_rows
                   if isinstance(row, dict) and row.get("symbol")]
        if len(tickers) >= 500:
            return tickers
    except Exception:
        pass

    # ── Source 3: SEC EDGAR company_tickers_exchange.json ─────────
    try:
        resp = requests.get(
            "https://www.sec.gov/files/company_tickers_exchange.json",
            headers={"User-Agent": "stockscreener/2.0 contact@example.com"},
            timeout=30,
        )
        resp.raise_for_status()
        data   = resp.json()
        fields = data.get("fields", [])
        rows   = data.get("data", [])
        exch_i = fields.index("exchange") if "exchange" in fields else 3
        tick_i = fields.index("ticker")   if "ticker"   in fields else 2
        target = "NYSE" if exchange == "nyse" else "Nasdaq"
        tickers = [
            row[tick_i].strip().replace(".", "-")
            for row in rows
            if len(row) > max(exch_i, tick_i)
            and str(row[exch_i]).strip().lower() == target.lower()
            and row[tick_i]
        ]
        return tickers
    except Exception:
        return []


@st.cache_data(ttl=86400)   # cache for 24 hours so repeated runs are instant
def load_tickers(exchange_key: str) -> list:
    """
    Return a deduplicated list of tickers for the chosen exchange.
    Results cached for 24 hours — clear cache if count looks wrong.
    """
    if exchange_key == "sp500":
        url  = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        df = pd.read_html(resp.text)[0]
        return df["Symbol"].str.replace(".", "-", regex=False).tolist()

    elif exchange_key in ("nyse", "nasdaq"):
        tickers = _fetch_exchange_tickers(exchange_key)
        # Deduplicate while preserving order
        seen = set()
        unique = []
        for t in tickers:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        return unique

    return []


def is_etf_or_fund(info: dict) -> bool:
    """
    Return True if the security is an ETF, index fund, trust, or other
    non-single-company product that should be excluded from the screener.

    Priority order:
    1. yfinance quoteType (most reliable — "etf", "mutualfund" etc.)
    2. Name keyword matching (conservative list — no false positives on real companies)
    """
    # quoteType is the most reliable signal — yfinance classifies it directly
    quote_type = (info.get("quoteType") or "").lower()
    if quote_type in ("etf", "mutualfund", "index", "future", "option",
                      "currency", "cryptocurrency"):
        return True

    # EQUITY = definitely a company stock — never exclude based on name alone
    if quote_type == "equity":
        return False

    # For anything else (or missing quoteType), check name keywords
    # Use only unambiguous ETF brand names — NOT generic English words
    name = (info.get("longName") or info.get("shortName") or "").lower()
    return any(kw in name for kw in ETF_KEYWORDS)

def build_excel(display: pd.DataFrame) -> bytes:
    """Write the display DataFrame (already filtered to visible columns) to Excel.
    Strips any remaining internal columns just in case."""
    # Belt-and-suspenders: drop internal/hidden columns that should never appear
    INTERNAL_COLS = {"_hist", "RangeHigh", "RangeLow", "RangePos",
                     "ROIC", "OBV", "PCV", "RSI", "MACD", "GoldenCross",
                     "MFI", "MFISweetSpot", "NoBearDiv"}
    df_out = display.drop(columns=[c for c in INTERNAL_COLS if c in display.columns],
                          errors="ignore")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Stock_Picks")
        sheet = writer.book["Stock_Picks"]
        col_map = {cell.value: get_column_letter(cell.column) for cell in sheet[1]}
        for col_name in ["RevenueGrowth", "EarningsGrowth"]:
            if col_name in col_map:
                for cell in sheet[col_map[col_name]][1:]:
                    if cell.value is not None:
                        cell.number_format = "0.00%"
        if "OwnerEarnings" in col_map:
            for cell in sheet[col_map["OwnerEarnings"]][1:]:
                if cell.value is not None:
                    cell.number_format = "$#,##0"
        for idx, col in enumerate(df_out.columns, 1):
            col_letter = get_column_letter(idx)
            max_len = max([len(str(c.value)) for c in sheet[col_letter] if c.value] + [len(col)])
            sheet.column_dimensions[col_letter].width = max_len + 2
    return output.getvalue()

# ───────────────────────────────────────────────────────────────
# SCORE COLOR MAP
# ───────────────────────────────────────────────────────────────

def color_score(val):
    try:
        v = float(val)
        if v >= 8:   return "background-color: #1a7a1a; color: white"
        elif v >= 5: return "background-color: #4caf50; color: white"
        elif v >= 3: return "background-color: #ff9800; color: white"
        else:        return "background-color: #f44336; color: white"
    except:
        return ""

# ───────────────────────────────────────────────────────────────
# MAIN RUN LOGIC
# ───────────────────────────────────────────────────────────────

with tab_screener:
 if run_btn:
    active_metrics = [k for k, v in metric_enabled.items() if v]

    exchange_key  = EXCHANGES[exchange]
    sector_label  = sector if sector != "All Sectors" else "All Sectors"

    # ── Check if we already have cached scan data for this exchange ──
    # Cache is valid as long as the exchange matches — mfi_period and range_days
    # are now recomputed client-side from stored raw history (zero re-downloads).
    cache         = st.session_state.get("screener_cache", {})
    cache_valid   = (
        cache.get("exchange_key") == exchange_key and
        bool(cache.get("results"))
    )

    if cache_valid:
        raw_results = cache["results"]
        # Recompute indicators with current sidebar settings from stored OHLCV —
        # this is pure CPU math, no network calls.
        results = recompute_indicators(raw_results, mfi_period, range_days)
        st.success(
            f"⚡ Using cached data from last scan "
            f"({len(results)} tickers · {cache.get('scanned_at','')}) — "
            f"indicators recomputed instantly from stored history. "
            f"Click **Clear Cache & Rescan** to fetch fresh data from the web."
        )
        if st.button("🔄 Clear Cache & Rescan", key="clear_cache_btn"):
            st.session_state.pop("screener_cache", None)
            st.rerun()
    else:
        # Load tickers for selected exchange (cached for 24hrs)
        with st.spinner(f"Loading {exchange} tickers..."):
            try:
                tickers = load_tickers(exchange_key)
            except Exception as e:
                st.error(f"Failed to load tickers: {e}")
                st.stop()

        if not tickers:
            st.error(f"No tickers returned for {exchange}. Try again later.")
            st.stop()

        # Warn if count looks suspiciously low (pagination may have failed)
        expected_min = {"sp500": 400, "nyse": 1500, "nasdaq": 2000}
        if len(tickers) < expected_min.get(exchange_key, 400):
            st.warning(
                f"⚠️ Only **{len(tickers)}** tickers loaded for {exchange} "
                f"(expected {expected_min.get(exchange_key, '?')}+). "
                f"The ticker source may be rate-limited — results will be incomplete. "
                f"Try clearing the 24hr cache and reloading."
            )

        st.info(
            f"Scanning **{len(tickers)}** tickers on **{exchange}** · "
            f"Sector: **{sector_label}** · ETFs/funds excluded · Est. time: ~5 min"
        )

        # ── Dynamic-worker scan ───────────────────────────────────────
        # ThreadPoolExecutor cannot resize mid-run, so we scan in batches.
        # After each batch we measure the pass rate and adjust worker count:
        #   pass rate ≥ 40% → healthy → ramp up (max 12)
        #   pass rate 20–40% → OK → hold
        #   pass rate 10–20% → slow → step down
        #   pass rate < 10%  → throttled → halve workers + pause
        #
        # Batch size scales with worker count so each batch takes ~10–15 s.
        # Status line updates after every batch with live worker count.

        progress_bar  = st.progress(0, text="Starting scan...")
        _conn_status  = st.empty()
        results       = []
        total         = len(tickers)
        done          = 0
        last_pct      = 0
        workers_now   = max_workers   # start from sidebar value
        remaining     = list(tickers)

        # Sliding window: track pass rate over last N tickers processed
        _window_done  = 0
        _window_pass  = 0
        _WINDOW       = 50            # evaluate every 50 tickers

        while remaining:
            # Batch size = workers × 8 so workers stay busy for several seconds
            batch_size  = max(workers_now * 8, 20)
            batch       = remaining[:batch_size]
            remaining   = remaining[batch_size:]

            with ThreadPoolExecutor(max_workers=workers_now) as executor:
                futures = {
                    executor.submit(process_ticker, (t, mfi_period, range_days)): t
                    for t in batch
                }
                for future in as_completed(futures):
                    result = future.result()
                    passed = result is not None
                    if passed:
                        results.append(result)
                    done          += 1
                    _window_done  += 1
                    _window_pass  += int(passed)

                    pct  = done / total
                    found = len(results)

                    if int(pct * 100) >= last_pct + 2 or done == total:
                        last_pct = int(pct * 100)
                        progress_bar.progress(
                            min(pct, 1.0),
                            text=(f"Scanning... {done}/{total} tickers  ·  "
                                  f"{found} returned data  ·  "
                                  f"{workers_now} workers  ({last_pct}%)")
                        )

            # ── Evaluate pass rate over last window and adjust workers ──
            if _window_done >= _WINDOW:
                pass_rate    = _window_pass / _window_done
                _window_done = 0
                _window_pass = 0

                if pass_rate >= 0.40:
                    # Healthy — ramp up (cap at 12 to stay safe on cloud)
                    new_workers = min(workers_now + 2, 12)
                    _icon = "🟢"
                    _msg  = f"healthy ({pass_rate:.0%}) → ramping up"
                elif pass_rate >= 0.20:
                    # Good enough — hold steady
                    new_workers = workers_now
                    _icon = "🟢"
                    _msg  = f"good ({pass_rate:.0%}) → holding at {workers_now}"
                elif pass_rate >= 0.10:
                    # Slowing — step down one
                    new_workers = max(workers_now - 1, 2)
                    _icon = "🟡"
                    _msg  = f"slowing ({pass_rate:.0%}) → stepping down"
                else:
                    # Throttled — halve and pause to let rate limit reset
                    new_workers = max(workers_now // 2, 1)
                    _icon = "🔴"
                    _msg  = f"throttled ({pass_rate:.0%}) → halving + pausing 15s"
                    time.sleep(15)

                workers_now = new_workers
                # Persist adjusted value so sidebar shows current count
                st.session_state["max_workers_val"] = workers_now
                _conn_status.caption(
                    f"{_icon} Workers: **{workers_now}** · {_msg} · "
                    f"Total found: {len(results)}"
                )

        progress_bar.empty()
        _conn_status.empty()

        if not results:
            # Run a quick health check to give specific advice
            with st.spinner("Diagnosing connection issue..."):
                _diag = check_yfinance_health()
            _diag_st = _diag["status"]

            if _diag_st in ("rate_limited", "slow"):
                st.error(
                    "🔴 **Rate Limited** — yfinance blocked the scan requests. "
                    "Too many parallel workers sent too many requests too fast.\n\n"
                    "**Fixes (try in order):**\n"
                    "1. Use the **Check Connection** button in the sidebar → click **Reduce Workers to 3**\n"
                    "2. Wait 60 seconds and run again\n"
                    "3. Switch to **S&P 500** (fewer tickers = fewer requests)"
                )
            elif _diag_st == "timeout":
                st.error(
                    "⏱️ **Connection Timed Out** — yfinance servers are not responding.\n\n"
                    "**Fixes:**\n"
                    "1. Wait 30–60 seconds and try again\n"
                    "2. Check your internet connection"
                )
            elif _diag_st == "blocked":
                st.error(
                    "⛔ **Connection Blocked** — the network cannot reach yfinance.\n\n"
                    "**Fixes:**\n"
                    "1. Switch to **S&P 500** only (smaller, more reliable)\n"
                    "2. If on Streamlit Cloud, the outbound network may have restrictions — "
                    "try running locally instead"
                )
            else:
                st.error(
                    "⚠️ **No data returned** — all tickers failed.\n\n"
                    "**Fixes:**\n"
                    "1. Reduce **Parallel Workers** to 3–5 in the sidebar\n"
                    "2. Wait 60 seconds for yfinance rate limits to reset\n"
                    "3. Try **S&P 500** instead of NYSE/NASDAQ"
                )
            st.session_state["_yf_health"] = _diag
            st.stop()

        # ── Store results in session_state cache ──────────────────
        # Raw OHLCV history is embedded in each result row (_hist key).
        # mfi_period and range_days are NOT stored — indicators are
        # recomputed on-the-fly from _hist whenever settings change.
        st.session_state["screener_cache"] = {
            "exchange_key": exchange_key,
            "results":      results,
            "scanned_at":   datetime.now().strftime("%H:%M:%S"),
        }
        # Recompute immediately with current sidebar values
        results = recompute_indicators(results, mfi_period, range_days)

    # Build DataFrame
    df = pd.DataFrame(results)
    df.replace(["N/A", "None", "-", ""], pd.NA, inplace=True)

    # ── Ensure ALL expected columns exist — fills missing ones with 0/NaN ──
    # This prevents KeyError crashes when cached data predates new metrics
    # or when recompute_indicators couldn't rebuild them (missing _hist).
    _zero_cols = ["OBV", "MFI", "PCV", "RSI", "MACD", "GoldenCross",
                  "MFISweetSpot", "NoBearDiv", "MA50Proximity",
                  "OE_Yield", "ROIC", "ROIC_Trend", "Piotroski",
                  "RevenueGrowth", "EarningsGrowth", "RangePct",
                  "RangePos", "RangeHigh", "RangeLow", "MA50"]
    for _c in _zero_cols:
        if _c not in df.columns:
            df[_c] = 0.0
    for _c in ["Price", "MarketCap", "P/E", "OwnerEarnings"]:
        if _c not in df.columns:
            df[_c] = np.nan
    for _c in ["Ticker", "Sector"]:
        if _c not in df.columns:
            df[_c] = ""

    numeric_cols = ["Price", "MA50", "RangeHigh", "RangeLow", "RangePct", "RangePos",
                    "MarketCap", "P/E", "OwnerEarnings", "OE_Yield",
                    "ROIC", "ROIC_Trend", "RevenueGrowth", "EarningsGrowth",
                    "Piotroski", "OBV", "MFI", "PCV",
                    "RSI", "MACD", "GoldenCross", "MFISweetSpot", "NoBearDiv", "MA50Proximity"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["ROIC_Trend"] = df["ROIC_Trend"].fillna(np.nan)
    for vol_col in ["OBV", "MFI", "PCV", "RSI", "MACD", "GoldenCross",
                    "MFISweetSpot", "NoBearDiv", "MA50Proximity"]:
        if vol_col in df.columns:
            df[vol_col] = df[vol_col].fillna(0)

    # ── Sector filter ──────────────────────────────────────────────
    if sector != "All Sectors":
        df = df[df["Sector"].str.strip().str.lower() == sector.strip().lower()]
        if df.empty:
            st.error(f"No results found for sector: **{sector}**. The sector name may differ from yfinance labels.")
            st.stop()

    # Safe round helper — only rounds if column exists and is numeric
    def _safe_round(col, decimals=2):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(decimals)

    _safe_round("P/E");  _safe_round("OE_Yield"); _safe_round("ROIC")
    _safe_round("ROIC_Trend"); _safe_round("OBV"); _safe_round("MFI")
    _safe_round("PCV"); _safe_round("MA50"); _safe_round("RangeHigh")
    _safe_round("RangeLow"); _safe_round("RangePct"); _safe_round("RangePos", 4)

    # MFI signal label
    def _mfi_signal(v):
        if pd.isna(v):   return "—"
        mfi = v * 100
        if mfi >= 80:    return "🔥 Overbought"
        if mfi >= 60:    return "📈 Buying"
        if mfi >= 40:    return "➡️ Neutral"
        if mfi >= 20:    return "📉 Selling"
        return "🧊 Oversold"
    df["MFI_Signal"] = df["MFI"].apply(_mfi_signal)

    # RangePosScore = 1 − RangePos
    df["RangePosScore"] = (1 - df["RangePos"].fillna(0.5)).round(4)

    # ── Dynamic score — safe column access, never KeyErrors ───────
    score = pd.Series(0.0, index=df.index)
    for key in METRICS:
        if metric_enabled.get(key, False):
            w = metric_weight.get(key, METRICS[key]["weight"])
            if key in df.columns:
                score += pd.to_numeric(df[key], errors="coerce").fillna(0) * w
            # If column is missing, contributes 0 — doesn't crash
    df["Score"] = score.round(2)

    # ── Apply filters — all use safe NaN-aware boolean masks ──────
    under_price = df["Price"].isna() | (df["Price"] <= max_price)

    if active_metrics:
        above_score = df["Score"] >= min_score
    else:
        above_score = pd.Series(True, index=df.index)

    ma50_col = pd.to_numeric(df.get("MA50", pd.Series(np.nan, index=df.index)), errors="coerce")
    if use_ma50_filter == "below":
        below_ma50 = df["Price"].isna() | ma50_col.isna() | (df["Price"] <= ma50_col)
    elif use_ma50_filter == "above":
        below_ma50 = df["Price"].isna() | ma50_col.isna() | (df["Price"] >= ma50_col)
    else:
        below_ma50 = pd.Series(True, index=df.index)

    if use_range_filter:
        rng_col = pd.to_numeric(df.get("RangePct", pd.Series(np.nan, index=df.index)), errors="coerce")
        in_range = rng_col.isna() | (rng_col <= range_max_pct)
    else:
        in_range = pd.Series(True, index=df.index)

    if use_pe_filter:
        pe_series = pd.to_numeric(df.get("P/E", pd.Series(np.nan, index=df.index)), errors="coerce")
        in_pe = pe_series.notna() & (pe_series >= pe_min) & (pe_series <= pe_max)
    else:
        in_pe = pd.Series(True, index=df.index)

    if use_rev_filter:
        rev_series = pd.to_numeric(df.get("RevenueGrowth", pd.Series(np.nan, index=df.index)), errors="coerce")
        in_rev = rev_series.notna() & (rev_series >= rev_min_idx / 100.0)
    else:
        in_rev = pd.Series(True, index=df.index)

    # Sort by Score when metrics are active, otherwise by RangePct (tightest range first)
    sort_col = "Score" if active_metrics else "RangePct"
    sort_asc  = not active_metrics   # ascending for RangePct (tighter = better), descending for Score

    screened = (
        df[under_price & below_ma50 & above_score & in_range & in_pe & in_rev]
        .sort_values(sort_col, ascending=sort_asc, na_position="last")
        .head(top_n)
        .reset_index(drop=True)
    )

    # ── Active metrics summary banner ────────────────────────────
    if active_metrics:
        active_labels = [METRICS[k]["label"] for k in active_metrics]
        st.success(f"**Active metrics:** {' · '.join(active_labels)}")
    else:
        st.info("**No metrics active** — results sorted by tightest price range. Score column will show 0.")

    # ── Results summary cards ────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Tickers Scanned", len(df))
    col2.metric("Passed Filters",  len(screened))
    col3.metric("Avg Score",       f"{screened['Score'].mean():.2f}" if not screened.empty else "—")
    col4.metric("Top Score",       f"{screened['Score'].max():.2f}"  if not screened.empty else "—")

    # ── Filter diagnostic panel — shows exactly why stocks are blocked ──
    if screened.empty and not df.empty:
        with st.expander("🔬 Filter Diagnostics — why are 0 stocks passing?", expanded=True):
            st.markdown("**Live filter counts** — how many stocks survive each gate:")
            diag_cols = st.columns(3)

            _up  = under_price.sum()
            _bm  = below_ma50.sum()
            _as  = above_score.sum()
            _ir  = in_range.sum()
            _ip  = in_pe.sum()
            _iv  = in_rev.sum()
            _tot = len(df)

            diag_cols[0].metric("Price filter",     f"{_up}/{_tot}",
                help=f"max_price={max_price}")
            diag_cols[1].metric("MA50 filter",      f"{_bm}/{_tot}",
                help=f"mode={use_ma50_filter}")
            diag_cols[2].metric("Score filter",     f"{_as}/{_tot}",
                help=f"min_score={min_score:.1f}")

            diag_cols2 = st.columns(3)
            diag_cols2[0].metric("Range filter",    f"{_ir}/{_tot}",
                help=f"enabled={use_range_filter}")
            diag_cols2[1].metric("P/E filter",      f"{_ip}/{_tot}",
                help=f"enabled={use_pe_filter}")
            diag_cols2[2].metric("Rev Growth filter", f"{_iv}/{_tot}",
                help=f"enabled={use_rev_filter}")

            st.markdown("**Combined pass rates** (each gate applied cumulatively):")
            cum1 = (under_price).sum()
            cum2 = (under_price & below_ma50).sum()
            cum3 = (under_price & below_ma50 & above_score).sum()
            cum4 = (under_price & below_ma50 & above_score & in_range).sum()
            cum5 = (under_price & below_ma50 & above_score & in_range & in_pe).sum()
            cum6 = (under_price & below_ma50 & above_score & in_range & in_pe & in_rev).sum()
            st.markdown(
                f"After price: **{cum1}** → "
                f"After MA50: **{cum2}** → "
                f"After score: **{cum3}** → "
                f"After range: **{cum4}** → "
                f"After P/E: **{cum5}** → "
                f"After rev growth: **{cum6}**"
            )

            st.markdown("**Sample of top 5 stocks and their scores:**")
            _sample = df.nlargest(5, "Score")[
                ["Ticker", "Price", "MA50", "Score"] +
                [k for k in active_metrics if k in df.columns][:4]
            ].copy()
            st.dataframe(_sample, use_container_width=True)

            st.markdown("**Current filter settings:**")
            st.json({
                "max_price":         max_price,
                "min_score":         min_score,
                "use_ma50_filter":   use_ma50_filter,
                "use_range_filter":  use_range_filter,
                "range_max_pct":     range_max_pct if use_range_filter else "off",
                "use_pe_filter":     use_pe_filter,
                "pe_range":          f"{pe_min}–{pe_max}" if use_pe_filter else "off",
                "use_rev_filter":    use_rev_filter,
                "active_metrics":    active_metrics,
                "scores_min":        float(df["Score"].min()),
                "scores_max":        float(df["Score"].max()),
                "scores_mean":       float(df["Score"].mean()),
                "price_min":         float(df["Price"].dropna().min()) if not df["Price"].dropna().empty else "N/A",
                "price_max":         float(df["Price"].dropna().max()) if not df["Price"].dropna().empty else "N/A",
                "ma50_above_count":  int((df["Price"] >= df["MA50"]).sum()),
                "ma50_below_count":  int((df["Price"] < df["MA50"]).sum()),
            })

    st.divider()

    if screened.empty:
        st.warning("No stocks passed the current filters. Try loosening the Score threshold, disabling the MA50 filter, or selecting a different sector.")
    else:
        # Format display columns
        display = screened.copy()
        display["MarketCap"]      = display["MarketCap"].apply(lambda x: f"${x:,.0f}" if pd.notnull(x) else "")
        display["OwnerEarnings"]  = display["OwnerEarnings"].apply(lambda x: f"${x:,.0f}" if pd.notnull(x) else "")
        display["RevenueGrowth"]  = display["RevenueGrowth"].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "")
        display["EarningsGrowth"] = display["EarningsGrowth"].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "")
        display["RangePct"]       = display["RangePct"].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else "")
        display["RangePos"]       = display["RangePos"].apply(
            lambda x: f"{'▁▂▃▄▅▆▇█'[min(int(x*8),7)] if pd.notnull(x) else '—'} {x:.0%}" if pd.notnull(x) else "—"
        )

        # Round all remaining numeric columns to 2 decimal places
        # Exclude: dollar columns, percentage columns, and range cols already formatted as strings
        skip_cols = {"Ticker", "Sector", "MarketCap", "OwnerEarnings",
                     "RevenueGrowth", "EarningsGrowth", "RangePct", "RangePos",
                     "MFI_Signal", "OE_Yield", "ROIC", "ROIC_Trend",
                     "RSI", "MACD", "GoldenCross", "MFISweetSpot", "NoBearDiv", "MA50Proximity"}
        for col in display.columns:
            if col not in skip_cols:
                if pd.api.types.is_numeric_dtype(display[col]):
                    display[col] = display[col].apply(lambda x: round(x, 2) if pd.notnull(x) else x)

        # Columns always hidden from display (still used in score calculations)
        # ROIC, OBV, PCV, RSI, MACD, GoldenCross: used in scoring, hidden from table
        # MFI: replaced by MFI_Signal label — hide numeric score
        ALWAYS_HIDDEN = {"ROIC", "OBV", "PCV", "RSI", "MACD", "GoldenCross", "MFI", "MFISweetSpot", "NoBearDiv"}

        # Hide columns for metrics that are toggled off
        all_metric_keys = list(METRICS.keys())
        hidden_cols = [k for k in all_metric_keys if not metric_enabled.get(k, True) and k in display.columns]
        # Add always-hidden cols
        hidden_cols += [c for c in ALWAYS_HIDDEN if c in display.columns and c not in hidden_cols]
        # MFI_Signal: show when MFI metric is toggled ON; hide when toggled off
        if not metric_enabled.get("MFI", True):
            if "MFI_Signal" in display.columns:
                hidden_cols.append("MFI_Signal")
        display = display.drop(columns=hidden_cols + ["_hist"], errors="ignore")

        styled = display.style.applymap(color_score, subset=["Score"])

        st.subheader(f"Top {len(screened)} Stocks — {sector_label}")
        st.dataframe(styled, use_container_width=True, height=600)

        # Download button
        st.divider()
        excel_bytes = build_excel(display)
        st.download_button(
            label="⬇️ Download Excel",
            data=excel_bytes,
            file_name=f"stock_screener_{sector.replace(' ', '_')}_{datetime.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

 else:
    st.info("👈 Configure your filters in the sidebar, then click **Run Screener** to start.")
    st.markdown("""
    ### How to use
    1. **Pick an exchange** — S&P 500 (~500 stocks, fast), NYSE or NASDAQ (2,000–3,500 stocks, 30–70 min)
    2. **Pick a sector** — scan one GICS sector or all sectors
    3. **Toggle metrics on/off** — disabled metrics are excluded from scoring AND hidden from results
    4. **Set your filters** — price cap, min score, MA50 toggle
    5. Hit **🚀 Run Screener**

    > ⚠️ **ETFs, index funds, trusts, and other non-company securities are automatically excluded** from all results.

    ### Metric weights (when enabled)
    | Metric | Default Weight | Ideal |
    |---|---|---|
    | OE Yield | ×3 | > 5% |
    | ROIC | ×2 | > 15% |
    | ROIC Trend | ×2 | Positive |
    | Revenue Growth | ×1 | > 5% |
    | Earnings Growth | ×1 | > 5% |
    | Piotroski Score | ×1 | 7–9 |
    | OBV (On-Balance Volume) | ×1 | 1.0 = rising accumulation |
    | MFI (Money Flow Index) | ×1 | > 0.5 = buying pressure |
    | PCV (Price-Confirmed Volume) | ×1 | > 0.5 = up-day volume dominant |

    ### Active filters
    - **Exchange** — defines the universe of tickers to scan
    - **Sector** — narrows results to one GICS sector (applied post-scan)
    - **Price cap** — stocks above your max price are excluded
    - **MA50 toggle** — when on, only stocks trading *below* their 50-day MA are shown
    - **Min score** — removes low-conviction picks from the table
    """)


# ───────────────────────────────────────────────────────────────
# ANALYZE TAB — single ticker deep dive
# ───────────────────────────────────────────────────────────────
with tab_analyze:
    st.subheader("🔍 Single Stock Analyzer")
    st.caption("Enter any ticker and select which metrics to run. Works on any stock — S&P 500, NYSE, NASDAQ, or international.")

    # ── Input row — ticker box + analyze button side by side ────
    _pending = st.session_state.pop("_pending_ticker", None)
    if _pending:
        st.session_state["analyze_ticker_input"] = _pending

    inp_col, btn_col = st.columns([3, 1])
    with inp_col:
        ticker_input = st.text_input(
            "Stock Ticker",
            placeholder="e.g. AAPL, MSFT, TSLA",
            key="analyze_ticker_input",
            help="Enter the ticker symbol exactly as it appears on the exchange.",
            label_visibility="collapsed",
        ).strip().upper()
    with btn_col:
        analyze_btn = st.button("🔬 Analyze", type="primary", use_container_width=True)

    # ── Collapsible metrics selector ─────────────────────────────
    with st.expander("⚙️ Select Metrics", expanded=False):
        analyze_metrics = {}
        m_cols = st.columns(2)
        for i, (key, cfg) in enumerate(METRICS.items()):
            with m_cols[i % 2]:
                analyze_metrics[key] = st.checkbox(
                    cfg["label"], value=True,
                    key=f"analyze_{key}", help=cfg["desc"],
                )

    # ── Search History ────────────────────────────────────────────
    _history = st.session_state.get("analyze_history", [])
    if _history:
        with st.expander(f"🕐 Recent Searches ({len(_history)})", expanded=False):
            _active = st.session_state.get("analyze_data", {}).get("ticker", "")
            for _h in _history:
                _is_active = _h["ticker"] == _active
                _border = "2px solid #4CAF50" if _is_active else "1px solid #444"
                _bg     = "rgba(76,175,80,0.08)" if _is_active else "transparent"
                # Use a button that spans full width — no nested columns
                _btn_label = f"{'▶ ' if _is_active else ''}{_h['ticker']}  ·  {_h['name'][:35]}{'…' if len(_h['name'])>35 else ''}"
                if st.button(_btn_label, key=f"hist_btn_{_h['ticker']}",
                             use_container_width=True,
                             help=f"Re-analyze {_h['ticker']}",
                             type="primary" if _is_active else "secondary"):
                    st.session_state["_pending_ticker"] = _h["ticker"]
                    st.session_state["_rerun_ticker"]   = _h["ticker"]
                    st.rerun()
            if st.button("🗑️ Clear History", use_container_width=True, key="clear_hist_btn"):
                st.session_state["analyze_history"] = []
                st.rerun()

    # ── On Analyze click: fetch only what is missing from cache ──
    _rerun_ticker = st.session_state.pop("_rerun_ticker", None)
    if _rerun_ticker:
        ticker_input  = _rerun_ticker
        analyze_btn   = True

    if analyze_btn and ticker_input:
        _screener_cache  = st.session_state.get("screener_cache", {})
        _cached_results  = _screener_cache.get("results", [])
        _cached_row      = next(
            (r for r in _cached_results if r.get("Ticker") == ticker_input), None
        )

        # Ensure fin cache dict exists and has an entry for this ticker
        if "analyze_fin_cache" not in st.session_state:
            st.session_state["analyze_fin_cache"] = {}
        _fin_store = st.session_state["analyze_fin_cache"]
        if ticker_input not in _fin_store:
            _fin_store[ticker_input] = {}          # initialise empty — avoids KeyError
        _fc = _fin_store[ticker_input]             # shorthand reference

        # ── Validator: is a cached value actually usable? ──────────
        def _ok(key):
            v = _fc.get(key)
            if v is None:                                return False
            if isinstance(v, pd.DataFrame) and v.empty:  return False
            if isinstance(v, dict) and not v:             return False  # only reject empty dicts
            return True

        need_info = not _ok("info")
        need_hist = not _ok("hist")
        need_fin  = not _ok("fin")
        need_bal  = not _ok("bal")
        need_cf   = not _ok("cf")
        need_raw  = not _ok("raw")

        anything_missing = any([need_info, need_hist, need_fin, need_bal, need_cf, need_raw])

        if anything_missing:
            missing_labels = [k for k, n in [
                ("info", need_info), ("price history", need_hist),
                ("income stmt", need_fin), ("balance sheet", need_bal),
                ("cash flow", need_cf), ("metrics", need_raw),
            ] if n]
            with st.spinner(f"Fetching for **{ticker_input}**: {', '.join(missing_labels)}..."):
                # Create ticker object once — used for all fetches below
                _stock = yf.Ticker(ticker_input)

                # ── info ─────────────────────────────────────────
                if need_info:
                    try:
                        _info = _stock.info or {}
                        # Store whatever we got — even a partial dict is useful
                        if _info:
                            _fc["info"] = _info
                    except Exception:
                        _info = _fc.get("info", {})
                else:
                    _info = _fc.get("info", {})

                # Patch info with screener row data for any gaps
                if _cached_row:
                    _patch = {
                        "sector":         _cached_row.get("Sector"),
                        "currentPrice":   _cached_row.get("Price"),
                        "marketCap":      _cached_row.get("MarketCap"),
                        "trailingPE":     _cached_row.get("P/E"),
                        "revenueGrowth":  _cached_row.get("RevenueGrowth"),
                        "earningsGrowth": _cached_row.get("EarningsGrowth"),
                    }
                    for k, v in _patch.items():
                        if v is not None and not _info.get(k):
                            _info[k] = v

                # Only abort if we have absolutely nothing — no info AND no screener row
                if not _info and not _cached_row:
                    # ── Smart typo suggestions ──────────────────
                    _suggestions = {
                        "APPL":  "AAPL (Apple)",
                        "GOOGL": "GOOGL ✓ — if this failed, try GOOG",
                        "GOOG":  "GOOG ✓ — if this failed, try GOOGL",
                        "AMZN":  "AMZN ✓ — double-check spelling",
                        "TSLA":  "TSLA ✓ — double-check spelling",
                        "MSFT":  "MSFT ✓ — double-check spelling",
                        "META":  "META (formerly FB)",
                        "FB":    "META (Facebook rebranded to META in 2021)",
                        "TWTR":  "TWTR was delisted — Twitter is now private (X)",
                        "TWITCH":"TWITCH is not public — owned by Amazon (AMZN)",
                        "BRK":   "BRK.A or BRK.B (Berkshire Hathaway)",
                        "BRKA":  "BRK-A",
                        "BRKB":  "BRK-B",
                        "NVIDA": "NVDA (Nvidia)",
                        "NVDIA": "NVDA (Nvidia)",
                        "NFLX":  "NFLX ✓ — double-check spelling",
                        "BABA":  "BABA ✓ — if this fails try 9988.HK (Hong Kong listing)",
                    }
                    _hint = _suggestions.get(ticker_input.upper())
                    _msg  = f"**`{ticker_input}`** was not found. "
                    if _hint:
                        _msg += f"\n\n💡 Did you mean **{_hint}**?"
                    else:
                        _msg += (
                            f"\n\n**Common causes:**\n"
                            f"- Typo in the ticker (e.g. APPL instead of AAPL)\n"
                            f"- The stock was delisted or went private\n"
                            f"- International stocks may need a suffix (e.g. `TSM` for TSMC, `ASML` for ASML)\n"
                            f"- yfinance may be rate-limited — wait 30 seconds and try again"
                        )
                    st.error(_msg)
                    st.stop()

                # ── price history ─────────────────────────────────
                if need_hist:
                    try:
                        _hist = _stock.history(period="1y")
                        if not _hist.empty:
                            _fc["hist"] = _hist
                    except Exception:
                        _hist = _fc.get("hist", pd.DataFrame())
                else:
                    _hist = _fc["hist"]

                # ── financial statements ──────────────────────────
                if need_fin:
                    try:
                        _fin = _stock.financials
                        _fc["fin"] = _fin
                    except Exception:
                        _fin = _fc.get("fin")
                else:
                    _fin = _fc.get("fin")

                if need_bal:
                    try:
                        _bal = _stock.balance_sheet
                        _fc["bal"] = _bal
                    except Exception:
                        _bal = _fc.get("bal")
                else:
                    _bal = _fc.get("bal")

                if need_cf:
                    try:
                        _cf = _stock.cashflow
                        _fc["cf"] = _cf
                    except Exception:
                        _cf = _fc.get("cf")
                else:
                    _cf = _fc.get("cf")

                # ── metrics (raw scores) ──────────────────────────
                if need_raw:
                    if _cached_row:
                        # Pull directly from screener data — no recalculation needed
                        _raw = {
                            "OE_Yield":       _cached_row.get("OE_Yield"),
                            "ROIC":           _cached_row.get("ROIC"),
                            "ROIC_Trend":     _cached_row.get("ROIC_Trend"),
                            "RevenueGrowth":  _cached_row.get("RevenueGrowth"),
                            "EarningsGrowth": _cached_row.get("EarningsGrowth"),
                            "Piotroski":      _cached_row.get("Piotroski"),
                            "OBV":            _cached_row.get("OBV"),
                            "MFI":            _cached_row.get("MFI"),
                            "PCV":            _cached_row.get("PCV"),
                        }
                    else:
                        try:
                            # Use pre-fetched statements — same pattern as screener
                            _fin_r = _fc.get("fin", pd.DataFrame())
                            _bal_r = _fc.get("bal", pd.DataFrame())
                            _cf_r  = _fc.get("cf",  pd.DataFrame())
                            _hist_r = _fc.get("hist", pd.DataFrame())
                            _oe, _oey = get_owner_earnings(_cf_r, _fin_r, _info)
                            _vols     = get_volume_signals(_hist_r, mfi_period)
                            _raw = {
                                "OE_Yield":       _oey,
                                "ROIC":           calculate_roic(_fin_r, _bal_r),
                                "ROIC_Trend":     calculate_roic_trend(_fin_r, _bal_r),
                                "RevenueGrowth":  _info.get("revenueGrowth"),
                                "EarningsGrowth": _info.get("earningsGrowth"),
                                "Piotroski":      calculate_piotroski(_fin_r, _bal_r, _cf_r),
                                "OBV":            _vols["OBV"],
                                "MFI":            _vols["MFI"],
                                "PCV":            _vols["PCV"],
                            }
                        except Exception:
                            _raw = _fc.get("raw", {})
                    _fc["raw"] = _raw
                else:
                    _raw = _fc["raw"]

        else:
            # ── Fully cached — zero API calls ────────────────────
            _info = _fc["info"]
            _hist = _fc["hist"]
            _fin  = _fc.get("fin")
            _bal  = _fc.get("bal")
            _cf   = _fc.get("cf")
            _raw  = _fc["raw"]

        # Persist everything back to store
        _fin_store[ticker_input].update({
            "info": _info, "hist": _hist,
            "fin":  _fin,  "bal":  _bal,
            "cf":   _cf,   "raw":  _raw,
        })

        st.session_state["analyze_data"] = {
            "ticker":      ticker_input,
            "info":        _info,
            "hist":        _hist,
            "fin":         _fin,
            "bal":         _bal,
            "cf":          _cf,
            "raw":         _raw,
            "metrics_sel": dict(analyze_metrics),
            "from_cache":  bool(_cached_row),
        }

        # ── Update search history ─────────────────────────────────
        _company_name = _info.get("longName") or _info.get("shortName") or ticker_input
        _hist_entry   = {"ticker": ticker_input, "name": _company_name}
        _history      = st.session_state.get("analyze_history", [])
        _history      = [h for h in _history if h["ticker"] != ticker_input]
        _history.insert(0, _hist_entry)
        st.session_state["analyze_history"] = _history[:30]

    elif analyze_btn and not ticker_input:
        st.warning("Please enter a ticker symbol.")

    st.divider()

    # ── Render from session_state — persists across all reruns ──
    if "analyze_data" not in st.session_state:
        st.info("Enter a ticker above and click **🔬 Analyze** to get started. Works on any publicly traded stock — S&P 500, NYSE, NASDAQ, or international.")
    else:
        try:
            _d       = st.session_state["analyze_data"]
            info     = _d["info"]
            hist_1y  = _d["hist"]
            fin_stmt = _d["fin"]
            bal_stmt = _d["bal"]
            cf_stmt  = _d["cf"]
            raw      = _d["raw"]
            sticker  = _d["ticker"]
            selected = [k for k, v in _d["metrics_sel"].items() if v]

            name    = info.get("longName") or info.get("shortName") or sticker
            price   = info.get("currentPrice") or info.get("regularMarketPrice")
            mktcap  = info.get("marketCap")

            st.markdown(f"## {name} &nbsp; `{sticker}`")
            # Show data source
            if _d.get("from_cache"):
                st.caption("⚡ Metrics from screener cache · Financials cached locally — no API calls")
            h1, h2, h3, h4 = st.columns(4)
            h1.metric("Price",    f"${price:,.2f}" if price else "N/A")
            h2.metric("Sector",   info.get("sector", "N/A"))
            h3.metric("Industry", info.get("industry", "N/A"))
            h4.metric("Mkt Cap",  f"${mktcap/1e9:.1f}B" if mktcap else "N/A")
            st.divider()

            if not selected:
                st.warning("No metrics were selected when Analyze was run.")
                st.stop()

            THRESHOLDS = {
                "OE_Yield":       (0.05,  0.02, True),
                "ROIC":           (0.15,  0.08, True),
                "ROIC_Trend":     (0.02,  0.0,  True),
                "RevenueGrowth":  (0.10,  0.05, True),
                "EarningsGrowth": (0.10,  0.05, True),
                "Piotroski":      (7,     4,    True),
                "OBV":            (0.8,   0.4,  True),
                "MFI":            (0.6,   0.4,  True),
                "PCV":            (0.5,   0.2,  True),
                "RSI":            (0.9,   0.5,  True),
                "MACD":           (0.9,   0.5,  True),
                "GoldenCross":    (0.9,   0.4,  True),
                "MFISweetSpot":   (0.9,   0.5,  True),
                "NoBearDiv":      (0.9,   0.4,  True),
                "MA50Proximity":  (0.9,   0.5,  True),
                "RangePosScore":  (0.75,  0.4,  True),
            }

            def _sig(key, val):
                if val is None or (isinstance(val, float) and np.isnan(val)): return "⚪"
                g, n, hi = THRESHOLDS.get(key, (None, None, True))
                if g is None: return "⚪"
                if hi:  return "🟢" if val >= g else ("🟡" if val >= n else "🔴")
                else:   return "🟢" if val <= g else ("🟡" if val <= n else "🔴")

            def _fv(key, val):
                if val is None or (isinstance(val, float) and np.isnan(val)): return "N/A"
                if key in ("OE_Yield","RevenueGrowth","EarningsGrowth","ROIC","ROIC_Trend"): return f"{val:.2%}"
                if key in ("OBV","MFI","PCV"): return f"{val:.4f}"
                if key == "Piotroski": return f"{int(val)} / 9"
                return str(round(val, 4))

            def _fb(val):
                if val is None or (isinstance(val, float) and np.isnan(val)): return "N/A"
                try:
                    v = float(val)
                    if abs(v)>=1e12: return f"${v/1e12:.2f}T"
                    if abs(v)>=1e9:  return f"${v/1e9:.2f}B"
                    if abs(v)>=1e6:  return f"${v/1e6:.2f}M"
                    if abs(v)>=1e3:  return f"${v/1e3:.2f}K"
                    return f"${v:,.2f}"
                except: return "N/A"

            def _fp(val):
                try:    return f"{float(val):.2%}" if val is not None else "N/A"
                except: return "N/A"

            def _fn(val, d=2):
                try:    return f"{float(val):,.{d}f}" if val is not None else "N/A"
                except: return "N/A"

            def irow(label, value, tip=None):
                # tip renders as a small ℹ tooltip after the label
                tip_html = (f" <span title='{tip}' style='cursor:help;color:#888;"
                            f"font-size:0.85em;'>ℹ️</span>") if tip else ""
                st.markdown(
                    f"<div style='display:flex;justify-content:space-between;"
                    f"padding:6px 0;border-bottom:1px solid #2a2a2a;'>"
                    f"<span style='color:#aaa;'>{label}{tip_html}</span>"
                    f"<span style='font-weight:600;'>{value}</span></div>",
                    unsafe_allow_html=True)

            # ── Metric cards ─────────────────────────────────
            st.markdown("### Metric Results")

            # Compute MFI signal for use in card and banner
            _mfi_raw = raw.get("MFI")
            def _mfi_label(v):
                if v is None or (isinstance(v, float) and np.isnan(v)): return None, None
                mfi = v * 100
                if mfi >= 80: return "🔥 Overbought",  "#ef5350"
                if mfi >= 60: return "📈 Buying",       "#26a69a"
                if mfi >= 40: return "➡️ Neutral",      "#888888"
                if mfi >= 20: return "📉 Selling",      "#f59e0b"
                return         "🧊 Oversold",           "#60a5fa"

            _mfi_lbl, _mfi_col = _mfi_label(_mfi_raw)

            # Show MFI signal banner if MFI is selected
            if "MFI" in selected and _mfi_lbl:
                _mfi_pct = f"{_mfi_raw*100:.1f}" if _mfi_raw is not None else "N/A"
                st.markdown(
                    f"<div style='background:{_mfi_col}22;border:1px solid {_mfi_col};"
                    f"border-radius:8px;padding:10px 16px;margin-bottom:12px;"
                    f"display:flex;justify-content:space-between;align-items:center;'>"
                    f"<span style='font-weight:600;font-size:1.05em;'>{_mfi_lbl}</span>"
                    f"<span style='color:#ccc;font-size:0.9em;'>MFI = {_mfi_pct} &nbsp;·&nbsp; "
                    f"Oversold ≤20 · Selling 20–40 · Neutral 40–60 · Buying 60–80 · Overbought ≥80</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

            cols = st.columns(3)
            for i, key in enumerate(selected):
                cfg = METRICS[key]; val = raw.get(key)
                # HTML-escape the description so special chars don't break the markup
                import html as _html
                safe_desc = _html.escape(cfg['desc'][:120])
                # Build extra badge line for MFI card
                extra = ""
                if key == "MFI" and _mfi_lbl:
                    extra = (f"<div style='margin-top:6px;padding:3px 8px;border-radius:4px;"
                             f"background:{_mfi_col}33;color:{_mfi_col};"
                             f"font-size:0.85em;font-weight:600;display:inline-block;'>"
                             f"{_mfi_lbl}</div>")
                with cols[i % 3]:
                    st.markdown(
                        f"<div style='border:1px solid #444;border-radius:8px;"
                        f"padding:14px;margin-bottom:12px;'>"
                        f"<div style='font-size:1.1em;font-weight:600;'>{_sig(key,val)} {cfg['label']}</div>"
                        f"<div style='font-size:1.8em;font-weight:700;margin:6px 0;'>{_fv(key,val)}</div>"
                        f"{extra}"
                        f"<div style='font-size:0.78em;color:#aaa;margin-top:6px;'>{safe_desc}...</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

            # ── MA50 ─────────────────────────────────────────
            st.divider()
            if not hist_1y.empty and len(hist_1y) >= 50:
                ma50_val = round(hist_1y["Close"].rolling(50).mean().iloc[-1], 2)
                diff_pct = ((price - ma50_val) / ma50_val * 100) if price else 0
                m1, m2, m3 = st.columns(3)
                m1.metric("Current Price", f"${price:,.2f}" if price else "N/A",
                          help="The latest traded price of the stock")
                m2.metric("50-Day MA",     f"${ma50_val:,.2f}",
                          help="50-Day Moving Average — the average closing price over the last 50 trading days. "
                               "Price below this line = stock is in a short-term downtrend or pullback")
                m3.metric("vs MA50",       f"{diff_pct:+.1f}%", delta_color="inverse",
                          help="How far the current price is above or below the 50-day MA. "
                               "Negative = trading below the MA (potential value zone). "
                               "Positive = trading above the MA (momentum zone)")
                st.caption("🔴 Above" if price > ma50_val else "🟢 Below" + " its 50-day moving average")

            # ── Price Range ───────────────────────────────────
            st.divider()
            st.markdown("### 📦 Price Range Analysis")
            st.caption(f"Last **{range_days} trading days** — adjust via sidebar")
            if not hist_1y.empty and len(hist_1y) >= range_days:
                win  = hist_1y["Close"].iloc[-range_days:]
                rh   = round(win.max(), 2); rl = round(win.min(), 2)
                mid  = (rh + rl) / 2
                rp   = round((rh - rl) / mid * 100, 2) if mid > 0 else None
                rpos = round((win.iloc[-1] - rl) / (rh - rl), 4) if (rh - rl) > 0 else 0.5
                rc1, rc2, rc3, rc4 = st.columns(4)
                rc1.metric("Range High", f"${rh:,.2f}",
                           help=f"The highest closing price over the last {range_days} trading days. "
                                "This acts as the top of the price range — a resistance level. "
                                "Price approaching this level may face selling pressure.")
                rc2.metric("Range Low",  f"${rl:,.2f}",
                           help=f"The lowest closing price over the last {range_days} trading days. "
                                "This acts as the bottom of the price range — a support level. "
                                "Price near this level may attract buyers.")
                rc3.metric("Range Width", f"{rp:.1f}%" if rp else "N/A",
                           help="Range Width = (High − Low) ÷ Midpoint × 100. "
                                "Measures how wide the trading channel is as a % of price. "
                                "Under 5% = very tight consolidation (coiling). "
                                "5–10% = normal consolidation. "
                                "Above 20% = high volatility / wide swings.")
                rc4.metric("Position",    f"{rpos:.0%}",
                           help="Where the current price sits within the range. "
                                "0% = at the range low (at support — potential buy zone). "
                                "50% = exactly at the midpoint. "
                                "100% = at the range high (at resistance — potential sell zone). "
                                "Stocks near the low end of a tight range can signal accumulation before a breakout.")
                bar = "█" * int(rpos*20) + "░" * (20 - int(rpos*20))
                pos_lbl = "Near Support 🟢" if rpos < 0.25 else ("Near Resistance 🔴" if rpos > 0.75 else "Mid-Range 🟡")
                st.markdown(
                    f"<div style='font-family:monospace;font-size:1.1em;margin:10px 0;'>"
                    f"${rl:,.2f} |{bar}| ${rh:,.2f}</div>"
                    f"<div style='color:#aaa;font-size:0.9em;'>Current ${price:,.2f} — {pos_lbl}</div>",
                    unsafe_allow_html=True)
                if rp is not None:
                    if rp <= 5:    st.success(f"🔒 Very Tight Range ({rp:.1f}%) — consolidation/breakout setup")
                    elif rp <= 10: st.info(f"📦 Tight Range ({rp:.1f}%) — consolidating")
                    elif rp <= 20: st.warning(f"📊 Moderate Range ({rp:.1f}%)")
                    else:          st.error(f"📉 Wide Range ({rp:.1f}%) — volatile")
                cdf = hist_1y["Close"].iloc[-range_days:].to_frame()
                cdf["High"] = rh; cdf["Low"] = rl
                st.line_chart(cdf, use_container_width=True)
            else:
                st.warning("Not enough price history to calculate range.")

            # ── Weighted score ────────────────────────────────
            st.divider()
            ts = sum(
                float(raw.get(k) or 0) * metric_weight.get(k, METRICS[k]["weight"])
                for k in selected
                if raw.get(k) is not None and not (isinstance(raw.get(k), float) and np.isnan(raw.get(k)))
            )
            st.metric("Weighted Score", f"{ts:.2f}")

            # ── Financial tabs ────────────────────────────────
            st.divider()
            st.markdown("## 📋 Full Financial Breakdown")
            ft = st.tabs(["🏢 Overview","💰 Valuation","📈 Income","🏦 Balance Sheet","💵 Cash Flow","📊 Price History"])

            with ft[0]:
                st.markdown("### 🏢 Company Overview")
                c1, c2 = st.columns(2)
                with c1:
                    irow("Full Name",    info.get("longName","N/A"))
                    irow("Exchange",     info.get("exchange","N/A"),   "The stock exchange where shares are listed and traded")
                    irow("Sector",       info.get("sector","N/A"),     "GICS sector classification — broad industry group the company belongs to")
                    irow("Industry",     info.get("industry","N/A"),   "Specific industry within the sector")
                    irow("Country",      info.get("country","N/A"),    "Country where the company is headquartered")
                    irow("Employees",    f"{info.get('fullTimeEmployees'):,}" if info.get("fullTimeEmployees") else "N/A",
                         "Total number of full-time employees")
                    irow("Website",      info.get("website","N/A"))
                with c2:
                    officers = info.get("companyOfficers",[])
                    irow("CEO",          officers[0].get("name","N/A") if officers else "N/A",
                         "Chief Executive Officer — the top executive responsible for running the company")
                    irow("Fiscal YE",    str(info.get("fiscalYearEnd","N/A")),
                         "Fiscal Year End — the month the company closes its annual accounting period")
                    irow("Audit Risk",   str(info.get("auditRisk","N/A")),
                         "Score 1–10 rating audit-related governance risk. Lower = less risk")
                    irow("Board Risk",   str(info.get("boardRisk","N/A")),
                         "Score 1–10 rating board structure and independence risk. Lower = less risk")
                    irow("Comp Risk",    str(info.get("compensationRisk","N/A")),
                         "Compensation Risk — score 1–10 rating executive pay structure risk. Lower = less risk")
                    irow("SH Rights",    str(info.get("shareHolderRightsRisk","N/A")),
                         "Shareholder Rights Risk — score 1–10. High = management has too much power vs shareholders")
                    irow("Overall Risk", str(info.get("overallRisk","N/A")),
                         "Overall governance risk score 1–10. Combines audit, board, compensation and shareholder rights scores")
                st.markdown("#### Business Summary")
                st.markdown(f"<div style='color:#ccc;line-height:1.6'>{info.get('longBusinessSummary','N/A')}</div>", unsafe_allow_html=True)

            with ft[1]:
                st.markdown("### 💰 Valuation")
                v1, v2 = st.columns(2)
                with v1:
                    irow("Market Cap",       _fb(info.get("marketCap")),
                         "Market Capitalization — total market value of all outstanding shares (Share Price × Shares Outstanding)")
                    irow("Enterprise Value", _fb(info.get("enterpriseValue")),
                         "Enterprise Value (EV) — Market Cap + Total Debt − Cash. Represents the true takeover cost of a company")
                    irow("Trailing P/E",     _fn(info.get("trailingPE")),
                         "Trailing Price-to-Earnings — share price divided by actual earnings per share over the last 12 months. Lower = cheaper relative to earnings")
                    irow("Forward P/E",      _fn(info.get("forwardPE")),
                         "Forward Price-to-Earnings — share price divided by analyst-estimated future earnings. Reflects growth expectations")
                    irow("PEG Ratio",        _fn(info.get("pegRatio")),
                         "Price/Earnings-to-Growth — P/E divided by earnings growth rate. Below 1.0 may indicate undervaluation relative to growth")
                    irow("Price/Sales",      _fn(info.get("priceToSalesTrailing12Months")),
                         "Price-to-Sales (P/S) — Market Cap divided by annual revenue. Useful for unprofitable companies. Lower = cheaper")
                    irow("Price/Book",       _fn(info.get("priceToBook")),
                         "Price-to-Book (P/B) — share price divided by book value per share (assets minus liabilities). Below 1.0 = trading below asset value")
                    irow("EV/Revenue",       _fn(info.get("enterpriseToRevenue")),
                         "Enterprise Value divided by annual revenue. Similar to P/S but accounts for debt. Lower = cheaper")
                    irow("EV/EBITDA",        _fn(info.get("enterpriseToEbitda")),
                         "Enterprise Value divided by EBITDA. Popular acquisition valuation multiple. Below 10 is generally considered reasonable")
                with v2:
                    irow("Trailing EPS",     _fn(info.get("trailingEps")),
                         "Trailing Earnings Per Share — actual net income per share over the last 12 months")
                    irow("Forward EPS",      _fn(info.get("forwardEps")),
                         "Forward Earnings Per Share — analyst-estimated earnings per share for the next 12 months")
                    irow("Book Value/Share", _fn(info.get("bookValue")),
                         "Net assets per share — total assets minus total liabilities, divided by shares outstanding")
                    irow("52-Wk High",       _fn(info.get("fiftyTwoWeekHigh")),
                         "Highest closing price over the last 52 weeks (one year)")
                    irow("52-Wk Low",        _fn(info.get("fiftyTwoWeekLow")),
                         "Lowest closing price over the last 52 weeks (one year)")
                    irow("50-Day MA",        _fn(info.get("fiftyDayAverage")),
                         "50-Day Moving Average — average closing price over the last 50 trading days. Used as a short-term trend indicator")
                    irow("200-Day MA",       _fn(info.get("twoHundredDayAverage")),
                         "200-Day Moving Average — average closing price over the last 200 trading days. Used as a long-term trend indicator")
                    irow("Beta",             _fn(info.get("beta")),
                         "Beta — measures stock volatility vs the S&P 500. Beta > 1 = more volatile than market; Beta < 1 = less volatile")
                    irow("Short % Float",    _fp(info.get("shortPercentOfFloat")),
                         "Short Interest as % of Float — percentage of tradeable shares currently sold short. High % can signal bearish sentiment or potential short squeeze")
                d1,d2,d3 = st.columns(3)
                d1.metric("Div Rate",  _fn(info.get("dividendRate")) if info.get("dividendRate") else "None",
                          help="Dividend Rate — annual cash dividend paid per share")
                d2.metric("Div Yield", _fp(info.get("dividendYield")) if info.get("dividendYield") else "None",
                          help="Dividend Yield — annual dividend as a % of share price. Higher = more income per dollar invested")
                d3.metric("Payout",    _fp(info.get("payoutRatio")) if info.get("payoutRatio") else "N/A",
                          help="Payout Ratio — percentage of net income paid out as dividends. Above 100% means paying more than it earns")

            with ft[2]:
                st.markdown("### 📈 Income Statement (Annual)")
                if fin_stmt is not None and not fin_stmt.empty:
                    fd = fin_stmt.T.copy()
                    fd.index = [str(i)[:10] for i in fd.index]
                    for c in fd.columns: fd[c] = fd[c].apply(lambda x: _fb(x) if pd.notnull(x) else "N/A")
                    st.dataframe(fd, use_container_width=True)
                else:
                    st.info("Income statement not available.")
                i1,i2,i3,i4 = st.columns(4)
                i1.metric("Revenue",    _fb(info.get("totalRevenue")),
                          help="Total Revenue — all money earned from selling products/services before any expenses are deducted")
                i2.metric("Gross",      _fb(info.get("grossProfits")),
                          help="Gross Profit — Revenue minus Cost of Goods Sold (COGS). Shows profit before operating expenses")
                i3.metric("EBITDA",     _fb(info.get("ebitda")),
                          help="Earnings Before Interest, Taxes, Depreciation & Amortization — a proxy for operating cash flow and profitability")
                i4.metric("Net Inc",    _fb(info.get("netIncomeToCommon")),
                          help="Net Income — the bottom line profit after all expenses, interest, and taxes have been deducted")
                i5,i6,i7,i8 = st.columns(4)
                i5.metric("Gross Mgn",  _fp(info.get("grossMargins")),
                          help="Gross Margin — Gross Profit ÷ Revenue. Higher % = more money left after production costs")
                i6.metric("Op Mgn",     _fp(info.get("operatingMargins")),
                          help="Operating Margin — Operating Income ÷ Revenue. Shows profit from core business operations")
                i7.metric("Net Mgn",    _fp(info.get("profitMargins")),
                          help="Net Profit Margin — Net Income ÷ Revenue. The % of every dollar of revenue that becomes profit")
                i8.metric("Rev Growth", _fp(info.get("revenueGrowth")),
                          help="Revenue Growth — year-over-year percentage increase in total revenue")

            with ft[3]:
                st.markdown("### 🏦 Balance Sheet (Annual)")
                if bal_stmt is not None and not bal_stmt.empty:
                    bd = bal_stmt.T.copy()
                    bd.index = [str(i)[:10] for i in bd.index]
                    for c in bd.columns: bd[c] = bd[c].apply(lambda x: _fb(x) if pd.notnull(x) else "N/A")
                    st.dataframe(bd, use_container_width=True)
                else:
                    st.info("Balance sheet not available.")
                b1,b2,b3,b4 = st.columns(4)
                b1.metric("Cash",      _fb(info.get("totalCash")),
                          help="Total Cash & Short-Term Investments — liquid assets the company can access immediately")
                b2.metric("Debt",      _fb(info.get("totalDebt")),
                          help="Total Debt — all short-term and long-term borrowings combined")
                b3.metric("Net Cash",  _fb((info.get("totalCash") or 0)-(info.get("totalDebt") or 0)),
                          help="Net Cash Position — Total Cash minus Total Debt. Positive = more cash than debt (strong balance sheet)")
                b4.metric("Assets",    _fb(info.get("totalAssets")),
                          help="Total Assets — everything the company owns: cash, property, equipment, intangibles, etc.")
                b5,b6,b7,b8 = st.columns(4)
                b5.metric("Cash/Shr",  _fn(info.get("totalCashPerShare")),
                          help="Cash Per Share — total cash divided by shares outstanding. Higher = more cash backing each share")
                b6.metric("D/E",       _fn(info.get("debtToEquity")),
                          help="Debt-to-Equity Ratio — total debt divided by shareholders equity. Lower = less financial leverage/risk. Above 2.0 can be concerning")
                b7.metric("Cur Ratio", _fn(info.get("currentRatio")),
                          help="Current Ratio — Current Assets ÷ Current Liabilities. Above 1.0 means the company can cover its short-term debts. Above 2.0 is considered healthy")
                b8.metric("Qck Ratio", _fn(info.get("quickRatio")),
                          help="Quick Ratio (Acid Test) — like Current Ratio but excludes inventory. Above 1.0 means the company can pay short-term debts without selling inventory")

            with ft[4]:
                st.markdown("### 💵 Cash Flow (Annual)")
                if cf_stmt is not None and not cf_stmt.empty:
                    cd = cf_stmt.T.copy()
                    cd.index = [str(i)[:10] for i in cd.index]
                    for c in cd.columns: cd[c] = cd[c].apply(lambda x: _fb(x) if pd.notnull(x) else "N/A")
                    st.dataframe(cd, use_container_width=True)
                else:
                    st.info("Cash flow not available.")
                cf1,cf2,cf3,cf4 = st.columns(4)
                cf1.metric("Op CF",   _fb(info.get("operatingCashflow")),
                           help="Operating Cash Flow — actual cash generated from core business operations. More reliable than net income as a profitability measure")
                cf2.metric("FCF",     _fb(info.get("freeCashflow")),
                           help="Free Cash Flow — Operating Cash Flow minus Capital Expenditures. Cash left over that can be used for dividends, buybacks, or debt repayment")
                cf3.metric("CapEx",   _fb(info.get("capitalExpenditures")),
                           help="Capital Expenditures — money spent on buying, maintaining or upgrading physical assets like buildings and equipment")
                sh = info.get("sharesOutstanding"); fcf = info.get("freeCashflow")
                cf4.metric("FCF/Shr", _fn(fcf/sh if fcf and sh else None),
                           help="Free Cash Flow Per Share — FCF divided by shares outstanding. How much free cash is generated per share owned")
                cf5,cf6 = st.columns(2)
                cf5.metric("ROA", _fp(info.get("returnOnAssets")),
                           help="Return on Assets — Net Income ÷ Total Assets. Measures how efficiently a company uses its assets to generate profit. Above 5% is generally good")
                cf6.metric("ROE", _fp(info.get("returnOnEquity")),
                           help="Return on Equity — Net Income ÷ Shareholders Equity. Measures how much profit is generated per dollar of shareholder investment. Above 15% is considered strong")

            with ft[5]:
                st.markdown("### 📊 Interactive Chart")

                # ── Chart controls row 1 ──────────────────────
                cc1, cc2, cc3, cc4 = st.columns(4)
                with cc1:
                    chart_type = st.selectbox(
                        "Chart Type",
                        ["Candlestick", "Heikin Ashi", "Line"],
                        index=0, key="chart_type",
                        help="Heikin Ashi smooths noise by averaging price data — easier to spot trends"
                    )
                with cc2:
                    interval = st.selectbox(
                        "Candle Interval",
                        ["1m","5m","15m","30m","1h","4h","1d","1wk","1mo"],
                        index=6, key="chart_interval",
                        help=(
                            "1m/5m: last 7 days max  |  "
                            "15m/30m/1h: last 60 days max  |  "
                            "4h: last 60 days max (resampled)  |  "
                            "1d/1wk/1mo: years of data available"
                        )
                    )
                with cc3:
                    # Restrict time period options based on interval
                    intraday_short = interval in ("1m", "5m")
                    intraday_mid   = interval in ("15m", "30m", "1h", "4h")

                    if intraday_short:
                        period_opts = ["1d", "5d", "7d"]
                        period_def  = 1
                    elif intraday_mid:
                        period_opts = ["1d", "5d", "1mo", "2mo"]
                        period_def  = 2
                    else:
                        period_opts = ["5d","1mo","3mo","6mo","1y","2y","5y","10y","Custom"]
                        period_def  = 4

                    pc = st.selectbox(
                        "Time Period", period_opts,
                        index=period_def, key="price_history_period"
                    )
                with cc4:
                    chart_theme = st.selectbox(
                        "Theme", ["Dark","Light"], index=0, key="chart_theme"
                    )

                # ── Custom date range (daily+ only) ───────────
                if pc == "Custom":
                    dr1, dr2 = st.columns(2)
                    with dr1:
                        custom_start = st.date_input(
                            "Start Date",
                            value=pd.Timestamp.today() - pd.Timedelta(days=365),
                            max_value=pd.Timestamp.today(),
                            key="chart_custom_start",
                        )
                    with dr2:
                        custom_end = st.date_input(
                            "End Date",
                            value=pd.Timestamp.today(),
                            max_value=pd.Timestamp.today(),
                            key="chart_custom_end",
                        )

                st.markdown("**Overlays**")
                ov1,ov2,ov3,ov4 = st.columns(4)
                show_ema20  = ov1.checkbox("EMA 20",  value=True,  key="show_ema20")
                show_ema50  = ov2.checkbox("EMA 50",  value=True,  key="show_ema50")
                show_ema200 = ov3.checkbox("EMA 200", value=False, key="show_ema200")
                show_bb     = ov4.checkbox("Bollinger Bands", value=False, key="show_bb")

                st.markdown("**Sub-Charts**")
                sc1,sc2,sc3 = st.columns(3)
                show_vol  = sc1.checkbox("Volume",  value=True,  key="show_vol")
                show_rsi  = sc2.checkbox("RSI",     value=True,  key="show_rsi")
                show_macd = sc3.checkbox("MACD",    value=False, key="show_macd")

                # ── Fetch history with correct interval ───────
                # yfinance interval/period compatibility:
                #   1m        → max period "7d"
                #   5m/15m/30m/1h → max period "60d"
                #   4h        → not native; fetch 1h and resample
                #   1d/1wk/1mo → any period
                fetch_interval = "1h" if interval == "4h" else interval

                # Map display period → yfinance period string for fetch
                period_fetch_map = {
                    "1d":"1d","5d":"5d","7d":"7d","1mo":"1mo","2mo":"2mo",
                    "3mo":"3mo","6mo":"6mo","1y":"2y","2y":"5y",
                    "5y":"10y","10y":"max","Custom":"max",
                }
                fetch_period = period_fetch_map.get(pc, "2y")

                try:
                    _s = yf.Ticker(sticker)
                    if pc == "Custom":
                        hist_full = _s.history(
                            start=str(custom_start),
                            end=str(custom_end),
                            interval=fetch_interval
                        )
                    else:
                        hist_full = _s.history(period=fetch_period, interval=fetch_interval)
                except Exception as _fe:
                    st.warning(f"Could not fetch {interval} data: {_fe}. Falling back to daily.")
                    hist_full = hist_1y

                # Resample 1h → 4h if needed
                if interval == "4h" and not hist_full.empty:
                    hist_full = hist_full.resample("4h").agg({
                        "Open":  "first",
                        "High":  "max",
                        "Low":   "min",
                        "Close": "last",
                        "Volume":"sum",
                    }).dropna()

                # Slice to display window for daily+ periods
                period_days_map = {
                    "5d":5,"1mo":21,"3mo":63,"6mo":126,
                    "1y":252,"2y":504,"5y":1260,"10y":2520,
                }
                if pc not in ("Custom","1d","5d","7d","1mo","2mo"):
                    dn = period_days_map.get(pc, 252)
                    hd = hist_full.iloc[-dn:].copy() if len(hist_full) >= dn else hist_full.copy()
                else:
                    hd = hist_full.copy()

                if not hd.empty:
                    import plotly.graph_objects as go
                    from plotly.subplots import make_subplots

                    bg   = "#0e1117" if chart_theme == "Dark" else "#ffffff"
                    fg   = "#ffffff" if chart_theme == "Dark" else "#000000"
                    grid = "#1f2937" if chart_theme == "Dark" else "#e5e7eb"

                    # ── Compute indicators ────────────────────
                    hd["EMA20"]  = hd["Close"].ewm(span=20,  adjust=False).mean()
                    hd["EMA50"]  = hd["Close"].ewm(span=50,  adjust=False).mean()
                    hd["EMA200"] = hd["Close"].ewm(span=200, adjust=False).mean()

                    # Bollinger Bands (20-day, 2σ)
                    bb_mid        = hd["Close"].rolling(20).mean()
                    bb_std        = hd["Close"].rolling(20).std()
                    hd["BB_mid"]  = bb_mid
                    hd["BB_up"]   = bb_mid + 2 * bb_std
                    hd["BB_low"]  = bb_mid - 2 * bb_std

                    # RSI (14-period)
                    delta  = hd["Close"].diff()
                    gain   = delta.clip(lower=0).rolling(14).mean()
                    loss   = (-delta.clip(upper=0)).rolling(14).mean()
                    rs     = gain / loss.replace(0, np.nan)
                    hd["RSI"] = 100 - (100 / (1 + rs))

                    # MACD (12/26/9)
                    ema12       = hd["Close"].ewm(span=12, adjust=False).mean()
                    ema26       = hd["Close"].ewm(span=26, adjust=False).mean()
                    hd["MACD"]  = ema12 - ema26
                    hd["Signal"]= hd["MACD"].ewm(span=9, adjust=False).mean()
                    hd["Hist"]  = hd["MACD"] - hd["Signal"]

                    # Heikin Ashi OHLC
                    ha = hd.copy()
                    ha["HA_Close"] = (hd["Open"] + hd["High"] + hd["Low"] + hd["Close"]) / 4
                    ha["HA_Open"]  = ((hd["Open"] + hd["Close"]) / 2).shift(1)
                    ha["HA_Open"].iloc[0] = (hd["Open"].iloc[0] + hd["Close"].iloc[0]) / 2
                    ha["HA_High"]  = pd.concat([hd["High"], ha["HA_Open"], ha["HA_Close"]], axis=1).max(axis=1)
                    ha["HA_Low"]   = pd.concat([hd["Low"],  ha["HA_Open"], ha["HA_Close"]], axis=1).min(axis=1)

                    # ── Build subplot layout ──────────────────
                    sub_charts = [s for s, show in [
                        ("Volume", show_vol), ("RSI", show_rsi), ("MACD", show_macd)
                    ] if show]

                    n_rows   = 1 + len(sub_charts)
                    row_h    = [0.55] + [0.45 / max(len(sub_charts), 1)] * len(sub_charts) if sub_charts else [1.0]
                    sub_specs= [[{"secondary_y": False}]] * n_rows

                    fig = make_subplots(
                        rows=n_rows, cols=1,
                        shared_xaxes=True,
                        row_heights=row_h,
                        vertical_spacing=0.03,
                        specs=sub_specs,
                    )

                    # ── Main price chart ──────────────────────
                    if chart_type == "Candlestick":
                        fig.add_trace(go.Candlestick(
                            x=hd.index, open=hd["Open"], high=hd["High"],
                            low=hd["Low"],  close=hd["Close"],
                            name="Price",
                            increasing_line_color="#26a69a",
                            decreasing_line_color="#ef5350",
                            increasing_fillcolor="#26a69a",
                            decreasing_fillcolor="#ef5350",
                        ), row=1, col=1)

                    elif chart_type == "Heikin Ashi":
                        fig.add_trace(go.Candlestick(
                            x=ha.index, open=ha["HA_Open"], high=ha["HA_High"],
                            low=ha["HA_Low"],  close=ha["HA_Close"],
                            name="Heikin Ashi",
                            increasing_line_color="#26a69a",
                            decreasing_line_color="#ef5350",
                            increasing_fillcolor="#26a69a",
                            decreasing_fillcolor="#ef5350",
                        ), row=1, col=1)

                    else:  # Line
                        fig.add_trace(go.Scatter(
                            x=hd.index, y=hd["Close"],
                            name="Close", line=dict(color="#2196f3", width=1.5)
                        ), row=1, col=1)

                    # ── Overlays ──────────────────────────────
                    if show_ema20:
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["EMA20"],
                            name="EMA 20", line=dict(color="#f59e0b", width=1),
                            opacity=0.85), row=1, col=1)
                    if show_ema50:
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["EMA50"],
                            name="EMA 50", line=dict(color="#a78bfa", width=1),
                            opacity=0.85), row=1, col=1)
                    if show_ema200:
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["EMA200"],
                            name="EMA 200", line=dict(color="#f87171", width=1),
                            opacity=0.85), row=1, col=1)
                    if show_bb:
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["BB_up"],
                            name="BB Upper", line=dict(color="#94a3b8", width=1, dash="dot"),
                            opacity=0.6), row=1, col=1)
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["BB_low"],
                            name="BB Lower", line=dict(color="#94a3b8", width=1, dash="dot"),
                            fill="tonexty", fillcolor="rgba(148,163,184,0.08)",
                            opacity=0.6), row=1, col=1)
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["BB_mid"],
                            name="BB Mid", line=dict(color="#94a3b8", width=1, dash="dash"),
                            opacity=0.4), row=1, col=1)

                    # ── Sub-charts ────────────────────────────
                    sub_row = 2
                    if show_vol:
                        colors = ["#26a69a" if hd["Close"].iloc[i] >= hd["Open"].iloc[i]
                                  else "#ef5350" for i in range(len(hd))]
                        fig.add_trace(go.Bar(
                            x=hd.index, y=hd["Volume"],
                            name="Volume", marker_color=colors, opacity=0.7,
                        ), row=sub_row, col=1)
                        fig.update_yaxes(title_text="Volume", row=sub_row, col=1,
                                         title_font=dict(size=10), tickfont=dict(size=9))
                        sub_row += 1

                    if show_rsi:
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["RSI"],
                            name="RSI", line=dict(color="#60a5fa", width=1.5)),
                            row=sub_row, col=1)
                        fig.add_hline(y=70, line_dash="dot", line_color="#ef5350",
                                      line_width=1, row=sub_row, col=1)
                        fig.add_hline(y=30, line_dash="dot", line_color="#26a69a",
                                      line_width=1, row=sub_row, col=1)
                        fig.update_yaxes(title_text="RSI", range=[0,100],
                                         row=sub_row, col=1,
                                         title_font=dict(size=10), tickfont=dict(size=9))
                        sub_row += 1

                    if show_macd:
                        hist_colors = ["#26a69a" if v >= 0 else "#ef5350"
                                       for v in hd["Hist"].fillna(0)]
                        fig.add_trace(go.Bar(x=hd.index, y=hd["Hist"],
                            name="MACD Hist", marker_color=hist_colors, opacity=0.6),
                            row=sub_row, col=1)
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["MACD"],
                            name="MACD", line=dict(color="#60a5fa", width=1.2)),
                            row=sub_row, col=1)
                        fig.add_trace(go.Scatter(x=hd.index, y=hd["Signal"],
                            name="Signal", line=dict(color="#f59e0b", width=1.2)),
                            row=sub_row, col=1)
                        fig.update_yaxes(title_text="MACD", row=sub_row, col=1,
                                         title_font=dict(size=10), tickfont=dict(size=9))

                    # ── Layout ────────────────────────────────
                    fig.update_layout(
                        height=600 + 120 * len(sub_charts),
                        paper_bgcolor=bg, plot_bgcolor=bg,
                        font=dict(color=fg, size=11),
                        legend=dict(
                            orientation="h", yanchor="bottom", y=1.01,
                            xanchor="left", x=0,
                            bgcolor="rgba(0,0,0,0)", font=dict(size=10)
                        ),
                        margin=dict(l=10, r=10, t=30, b=10),
                        xaxis_rangeslider_visible=False,
                        hovermode="x unified",
                    )
                    fig.update_xaxes(
                        gridcolor=grid, showgrid=True,
                        zeroline=False, showspikes=True,
                        spikecolor="#666", spikethickness=1,
                    )
                    fig.update_yaxes(
                        gridcolor=grid, showgrid=True,
                        zeroline=False,
                    )
                    fig.update_yaxes(title_text="Price ($)", row=1, col=1,
                                     title_font=dict(size=10), tickfont=dict(size=9))

                    st.plotly_chart(fig, use_container_width=True)

                    # ── Period stats ──────────────────────────
                    s1, s2, s3, s4 = st.columns(4)
                    pr = (hd["Close"].iloc[-1] - hd["Close"].iloc[0]) / hd["Close"].iloc[0]
                    s1.metric("Period Return", _fp(pr),     delta=f"{pr:+.2%}")
                    s2.metric("Period High",   f"${hd['High'].max():,.2f}")
                    s3.metric("Period Low",    f"${hd['Low'].min():,.2f}")
                    s4.metric("Avg Volume",    _fb(hd["Volume"].mean()).replace("$",""))
                else:
                    st.info("No price history available.")

        except Exception as e:
            st.error(f"Error rendering analysis: {e}")
